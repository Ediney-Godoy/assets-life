print("Main: Imports start", flush=True)
from fastapi import FastAPI, HTTPException, Depends, Request
print("Main: FastAPI imported", flush=True)
from fastapi.responses import JSONResponse
import re
from starlette.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from sqlalchemy.orm import Session
import sqlalchemy as sa
from datetime import date
import os

from .database import SessionLocal, engine
from .config import ALLOW_DDL, SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_USE_TLS, SMTP_USE_SSL, MAIL_FROM, MAIL_SENDER_NAME, SECRET_KEY, ALGORITHM, JWT_EXPIRE_MINUTES, RESET_TOKEN_EXPIRE_MINUTES
from .dependencies import get_db, get_current_user, get_allowed_company_ids, check_permission
from .models import Base as SA_Base, Company as CompanyModel
from .models import Employee as EmployeeModel, Vinculo as VinculoEnum, Status as StatusEnum
from .models import ManagementUnit as UGModel
from .models import Usuario as UsuarioModel
import bcrypt
from .models import RevisaoPeriodo as RevisaoPeriodoModel
from .models import RevisaoItem as RevisaoItemModel
from .models import RevisaoDelegacao as RevisaoDelegacaoModel
from .models import CentroCusto as CentroCustoModel
from .models import ClasseContabil as ClasseContabilModel
from .models import ContaContabil as ContaContabilModel
from .models import (
    GrupoPermissao as GrupoPermissaoModel,
    Transacao as TransacaoModel,
    GrupoEmpresa as GrupoEmpresaModel,
    GrupoTransacao as GrupoTransacaoModel,
    GrupoUsuario as GrupoUsuarioModel,
    AuditoriaLog as AuditoriaLogModel,
)
from .models import TokenRedefinicao as TokenRedefinicaoModel
from .models import Cronograma as CronogramaModel, CronogramaTarefa as CronogramaTarefaModel, CronogramaTarefaEvidencia as CronogramaTarefaEvidenciaModel
from fastapi import UploadFile, File
from datetime import datetime, timedelta
from jose import jwt, JWTError
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import secrets
print("Main: Importing routes...", flush=True)
from .routes.relatorios_rvu import router as relatorios_rvu_router
from .routes.supervisao_rvu import router as supervisao_rvu_router
from .routes.assets import router as assets_router
from .routes.reviews import router as reviews_router
print("Main: Routes imported", flush=True)
app = FastAPI(title="Asset Life API", version="0.2.0")
print("Main: FastAPI app created", flush=True)
print("Main: Including relatorios router...", flush=True)
app.include_router(relatorios_rvu_router)
print("Main: Including supervisao router...", flush=True)
app.include_router(supervisao_rvu_router)
print("Main: Including assets router...", flush=True)
app.include_router(assets_router)
print("Main: Including reviews router...", flush=True)
app.include_router(reviews_router)
print("Main: Routers included", flush=True)

print("Main: Creating tables...", flush=True)
# Create tables if they don't exist
# SA_Base.metadata.create_all(bind=engine)
print("Main: Tables created (SKIPPED)", flush=True)

# CORS de desenvolvimento: amplia suporte para localhost, 127.0.0.1 e IPs de rede
# Inclui origens comuns de Vite e permite configurar um origin adicional via env FRONTEND_ORIGIN
origins = [
    # localhost (Vite ports mais comuns)
    "http://localhost:5173", "http://localhost:5174", "http://localhost:5175", "http://localhost:5176",
    "http://localhost:5180", "http://localhost:5181",
    # 127.0.0.1 equivalentes
    "http://127.0.0.1:5173", "http://127.0.0.1:5174", "http://127.0.0.1:5175", "http://127.0.0.1:5176",
    "http://127.0.0.1:5180", "http://127.0.0.1:5181",
    # Produção - Frontend Vercel
    "https://assets-life-bp3b.vercel.app",
    # Backend Koyeb (para testes diretos)
    "https://brief-grete-assetlife-f50c6bd0.koyeb.app",
    "https://different-marlie-assetslifev2-bc199b4b.koyeb.app",
]

# Origin adicional via variável de ambiente (ex.: "http://192.168.101.8:5175" ou URL de produção)
_frontend_origin = os.getenv("FRONTEND_ORIGIN")
if _frontend_origin:
    # Permite múltiplas origens separadas por vírgula
    for origin in _frontend_origin.split(","):
        origin = origin.strip()
        if origin and origin not in origins:
            origins.append(origin)

# Relaxar regex para aceitar qualquer origem válida (evita bloqueios quando o proxy/ingress altera o Host/Origin)
ORIGIN_REGEX = r"^https?://.+$"
origin_re = re.compile(ORIGIN_REGEX)

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_origin_regex=ORIGIN_REGEX,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    allow_credentials=True,
    expose_headers=["Content-Disposition", "Content-Type"],
)

# Complemento CORS para "Private Network Access" (PNA) em navegadores modernos.
# Alguns ambientes dev via IP (ex.: 192.168.x.x) disparam preflight com
# "access-control-request-private-network: true" e exigem o cabeçalho abaixo.
# Este middleware adiciona o cabeçalho sem interferir no CORSMiddleware padrão.
from fastapi import Request

@app.middleware("http")
async def allow_private_network(request: Request, call_next):
    # Pre-attach permissive CORS headers for known/regex-matched origins to avoid race conditions
    origin = request.headers.get("origin")
    response = await call_next(request)
    try:
        if request.headers.get("access-control-request-private-network") == "true":
            response.headers["Access-Control-Allow-Private-Network"] = "true"
        if origin and (origin in origins or origin_re.match(origin)):
            if not response.headers.get("Access-Control-Allow-Origin"):
                response.headers["Access-Control-Allow-Origin"] = origin
                response.headers["Vary"] = "Origin"
            if not response.headers.get("Access-Control-Allow-Credentials"):
                response.headers["Access-Control-Allow-Credentials"] = "true"
            if not response.headers.get("Access-Control-Expose-Headers"):
                response.headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
            if not response.headers.get("Access-Control-Allow-Methods"):
                response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            if not response.headers.get("Access-Control-Allow-Headers"):
                req_headers = request.headers.get("access-control-request-headers", "*")
                response.headers["Access-Control-Allow-Headers"] = req_headers
    except Exception:
        pass
    return response

@app.exception_handler(HTTPException)
async def _http_exception_handler(request: Request, exc: HTTPException):
    origin = request.headers.get("origin")
    headers = {}
    try:
        if origin and (origin in origins or origin_re.match(origin)):
            headers["Access-Control-Allow-Origin"] = origin
            headers["Vary"] = "Origin"
            headers["Access-Control-Allow-Credentials"] = "true"
            headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
            headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            headers["Access-Control-Allow-Headers"] = request.headers.get("access-control-request-headers", "*")
    except Exception:
        pass
    return JSONResponse(status_code=exc.status_code, content={"detail": getattr(exc, "detail", "Erro")}, headers=headers)

@app.exception_handler(Exception)
async def _generic_exception_handler(request: Request, exc: Exception):
    origin = request.headers.get("origin")
    headers = {}
    try:
        if origin and (origin in origins or origin_re.match(origin)):
            headers["Access-Control-Allow-Origin"] = origin
            headers["Vary"] = "Origin"
            headers["Access-Control-Allow-Credentials"] = "true"
            headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
            headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            headers["Access-Control-Allow-Headers"] = request.headers.get("access-control-request-headers", "*")
    except Exception:
        pass
    return JSONResponse(status_code=500, content={"detail": "Erro interno"}, headers=headers)

@app.options("/{path:path}")
async def cors_preflight(request: Request, path: str):
    origin = request.headers.get("origin")
    headers_req = request.headers.get("access-control-request-headers", "*")
    allow = origin and (origin in origins or origin_re.match(origin))
    headers = {}
    if allow:
        headers = {
            "Access-Control-Allow-Origin": origin,
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, PATCH",
            "Access-Control-Allow-Headers": headers_req,
            "Access-Control-Allow-Credentials": "true",
            "Vary": "Origin",
            "Access-Control-Max-Age": "3600",
        }
    return Response(status_code=200, headers=headers)

# -----------------------------
# Auth (JWT) e Segurança
# -----------------------------
# Use chave vinda do ambiente em produção; mantém fallback dev para ambiente local
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-key-change-in-prod")
ALGORITHM = "HS256"
security = HTTPBearer()
JWT_EXPIRE_MINUTES = 60 * 8  # 8 horas
RESET_TOKEN_EXPIRE_MINUTES = 30  # 30 minutos

# Tentativas de login in-memory (substituir por campos na tabela usuarios posteriormente)
login_attempts: dict[str, dict] = {}

class LoginPayload(BaseModel):
    email: Optional[EmailStr] = None
    identificador: Optional[str] = None  # email ou nome de usuário
    senha: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    # first_access: bool = False

def create_access_token(data: dict, expires_minutes: int = 60):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=expires_minutes)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
 
# Dependencies are now imported from .dependencies

@app.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginPayload, request: Request, db: Session = Depends(get_db)):
    identifier = (payload.identificador or (payload.email or "")).strip()
    if not identifier:
        raise HTTPException(status_code=400, detail="Informe email ou usuário")

    key = identifier.lower()
    la = login_attempts.get(key)
    now = datetime.utcnow()
    if la and la.get("blocked_until") and la["blocked_until"] > now:
        raise HTTPException(status_code=403, detail="Usuário temporariamente bloqueado. Tente novamente mais tarde.")

    # Identificar por email ou nome de usuário
    if "@" in identifier:
        user = db.query(UsuarioModel).filter(UsuarioModel.email == identifier).first()
    else:
        user = db.query(UsuarioModel).filter(UsuarioModel.nome_usuario == identifier).first()

    client_ip = request.client.host if request.client else None

    if not user or not bcrypt.checkpw(payload.senha.encode("utf-8"), (user.senha_hash or "").encode("utf-8")):
        # Atualizar tentativas
        if not la:
            la = {"count": 0, "blocked_until": None}
        la["count"] = la.get("count", 0) + 1
        if la["count"] >= 3:
            la["blocked_until"] = now + timedelta(minutes=15)
        login_attempts[key] = la

        # Auditoria (falha)
        try:
            audit(db, usuario_id=user.id if user else None, acao="LOGIN_FALHA", entidade="Usuario", entidade_id=user.id if user else None, detalhes=f"ip={client_ip}")
        except Exception:
            pass
        raise HTTPException(status_code=401, detail="Credenciais inválidas")

    # Sucesso -> reset tentativas
    login_attempts.pop(key, None)

    token = create_access_token({"sub": str(user.id)}, expires_minutes=JWT_EXPIRE_MINUTES)

    # Auditoria (sucesso)
    try:
        audit(db, usuario_id=user.id, acao="LOGIN_SUCESSO", entidade="Usuario", entidade_id=user.id, detalhes=f"ip={client_ip}")
    except Exception:
        pass
    return TokenResponse(access_token=token)


class ForgotPasswordRequest(BaseModel):
    email: EmailStr


class ResetPasswordRequest(BaseModel):
    token: str
    nova_senha: str
    confirmar_senha: str


class ChangePasswordRequest(BaseModel):
    senha_atual: str
    nova_senha: str
    confirmar_senha: str


def validate_password_strength(p: str):
    if len(p) < 8:
        raise HTTPException(status_code=400, detail="Senha deve ter pelo menos 8 caracteres")
    # políticas mínimas: número, maiúscula e símbolo
    if not any(c.isupper() for c in p):
        raise HTTPException(status_code=400, detail="Senha deve conter pelo menos uma letra maiúscula")
    if not any(c.isdigit() for c in p):
        raise HTTPException(status_code=400, detail="Senha deve conter pelo menos um número")
    if not any(not c.isalnum() for c in p):
        raise HTTPException(status_code=400, detail="Senha deve conter pelo menos um símbolo")


from email.message import EmailMessage
import smtplib
import ssl

def _send_email(to: str, subject: str, body: str):
    if not SMTP_HOST:
        return False
    msg = EmailMessage()
    msg["From"] = f"{MAIL_SENDER_NAME} <{MAIL_FROM}>"
    msg["To"] = to
    msg["Subject"] = subject
    msg.set_content(body)
    try:
        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
                if SMTP_USER:
                    server.login(SMTP_USER, SMTP_PASSWORD or "")
                server.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                if SMTP_USE_TLS:
                    server.starttls(context=ssl.create_default_context())
                if SMTP_USER:
                    server.login(SMTP_USER, SMTP_PASSWORD or "")
                server.send_message(msg)
        return True
    except Exception:
        return False


@app.post("/auth/forgot-password")
def forgot_password(payload: ForgotPasswordRequest, db: Session = Depends(get_db)):
    user = db.query(UsuarioModel).filter(UsuarioModel.email == payload.email).first()
    # Não revelar existência do email
    if user:
        # invalidar tokens antigos não usados?
        token = secrets.token_urlsafe(48)
        expires = datetime.utcnow() + timedelta(minutes=RESET_TOKEN_EXPIRE_MINUTES)
        db.add(TokenRedefinicaoModel(usuario_id=user.id, token=token, expiracao=expires, usado=False))
        db.commit()
        try:
            base_url = os.getenv("FRONTEND_BASE_URL", "http://localhost:5173")
            link = f"{base_url}/reset-password?token={token}"
            _send_email(
                to=user.email,
                subject="Redefinição de Senha - Asset Life",
                body=f"Olá {user.nome_completo},\n\nRecebemos uma solicitação para redefinir sua senha. Use o link abaixo (válido por 30 minutos):\n{link}\n\nSe você não solicitou, ignore este email.",
            )
            audit(db, usuario_id=user.id, acao="RESET_SOLICITADO", entidade="Usuario", entidade_id=user.id)
        except Exception:
            pass
    return {"ok": True}


@app.post("/auth/reset-password")
def reset_password(payload: ResetPasswordRequest, db: Session = Depends(get_db)):
    if payload.nova_senha != payload.confirmar_senha:
        raise HTTPException(status_code=400, detail="Confirmação de senha não confere")
    validate_password_strength(payload.nova_senha)
    t = db.query(TokenRedefinicaoModel).filter(TokenRedefinicaoModel.token == payload.token).first()
    if not t or t.usado:
        raise HTTPException(status_code=400, detail="Token inválido")
    if t.expiracao < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Token expirado")
    user = db.query(UsuarioModel).filter(UsuarioModel.id == t.usuario_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    # atualizar senha
    new_hash = bcrypt.hashpw(payload.nova_senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    user.senha_hash = new_hash
    t.usado = True
    db.commit()
    try:
        audit(db, usuario_id=user.id, acao="RESET_CONCLUIDO", entidade="Usuario", entidade_id=user.id)
    except Exception:
        pass
    return {"ok": True}


@app.post("/auth/change-password")
def change_password(payload: ChangePasswordRequest, current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if payload.nova_senha != payload.confirmar_senha:
        raise HTTPException(status_code=400, detail="Confirmação de senha não confere")
    if not bcrypt.checkpw(payload.senha_atual.encode("utf-8"), (current_user.senha_hash or "").encode("utf-8")):
        raise HTTPException(status_code=400, detail="Senha atual incorreta")
    validate_password_strength(payload.nova_senha)
    new_hash = bcrypt.hashpw(payload.nova_senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    current_user.senha_hash = new_hash
    db.commit()
    try:
        audit(db, usuario_id=current_user.id, acao="ALTERACAO_SENHA", entidade="Usuario", entidade_id=current_user.id)
    except Exception:
        pass
    return {"ok": True}

@app.get("/auth/me", response_model=dict)
def auth_me(current_user: UsuarioModel = Depends(get_current_user)):
    return {"id": current_user.id, "nome": current_user.nome_completo, "email": current_user.email}

@app.get("/auth/me/permissoes", response_model=dict)
def auth_me_permissions(current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    # grupos do usuário
    links = db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.usuario_id == current_user.id).all()
    grupo_ids = [l.grupo_id for l in links]
    grupos = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id.in_(grupo_ids)).all()
    # transações liberadas
    trans_links = db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id.in_(grupo_ids)).all()
    trans_ids = [t.transacao_id for t in trans_links]
    transacoes = db.query(TransacaoModel).filter(TransacaoModel.id.in_(trans_ids)).all()
    rotas = sorted({t.rota for t in transacoes})
    # empresas liberadas
    emp_links = db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id.in_(grupo_ids)).all()
    empresas_ids = sorted({e.empresa_id for e in emp_links})
    # Fallbacks
    if not empresas_ids and current_user.empresa_id is not None:
        empresas_ids = [current_user.empresa_id]
    if not rotas:
        rotas = ["/dashboard"]
    return {
        "usuario_id": current_user.id,
        "grupos": [{"id": g.id, "nome": g.nome} for g in grupos],
        "empresas_ids": empresas_ids,
        "rotas": rotas,
    }

 # (get_db moved above to avoid NameError during Depends evaluation)

# @app.on_event("startup")
def on_startup():
    # Ensure tables exist for development environments
    if ALLOW_DDL:
        try:
            SA_Base.metadata.create_all(bind=engine)
        except Exception as e:
            import traceback
            print("DB init error:", e)
            traceback.print_exc()
        # Ajustes de schema leves (sem Alembic): adiciona coluna 'tipo' em cronogramas_tarefas se ausente
        try:
            with engine.connect() as conn:
                conn.execute(sa.text("ALTER TABLE cronogramas_tarefas ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'Tarefa'"))
                conn.execute(sa.text(
                    """
                    CREATE TABLE IF NOT EXISTS cronogramas_tarefas_evidencias (
                        id SERIAL PRIMARY KEY,
                        tarefa_id INTEGER NOT NULL REFERENCES cronogramas_tarefas(id),
                        nome_arquivo VARCHAR(255) NOT NULL,
                        content_type VARCHAR(100) NOT NULL,
                        tamanho_bytes INTEGER NOT NULL,
                        conteudo BYTEA NOT NULL,
                        criado_em TIMESTAMP DEFAULT NOW() NOT NULL,
                        uploaded_by INTEGER NULL REFERENCES usuarios(id)
                    )
                    """
                ))
                conn.execute(sa.text("CREATE UNIQUE INDEX IF NOT EXISTS ux_cronogramas_periodo ON cronogramas(periodo_id)"))
        except Exception as e:
            # Não bloquear startup por falhas de ALTER
            import traceback
            print("Schema tweak error (tipo coluna):", e)
            traceback.print_exc()

    # Executa ajustes mínimos de schema mesmo quando ALLOW_DDL=false (idempotentes e seguros)
    try:
        with engine.connect() as conn:
            conn.execute(sa.text("ALTER TABLE cronogramas_tarefas ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'Tarefa'"))
            conn.execute(sa.text(
                """
                CREATE TABLE IF NOT EXISTS cronogramas_tarefas_evidencias (
                    id SERIAL PRIMARY KEY,
                    tarefa_id INTEGER NOT NULL REFERENCES cronogramas_tarefas(id),
                    nome_arquivo VARCHAR(255) NOT NULL,
                    content_type VARCHAR(100) NOT NULL,
                    tamanho_bytes INTEGER NOT NULL,
                    conteudo BYTEA NOT NULL,
                    criado_em TIMESTAMP DEFAULT NOW() NOT NULL,
                    uploaded_by INTEGER NULL REFERENCES usuarios(id)
                )
                """
            ))
            conn.execute(sa.text("CREATE UNIQUE INDEX IF NOT EXISTS ux_cronogramas_periodo ON cronogramas(periodo_id)"))
        print("Main: Schema adjusted", flush=True)
    except Exception as e:
        import traceback
        print("Schema tweak error (tipo/evidencias fallback):", e)
        traceback.print_exc()

    # Seed default transactions if none exist or missing
    try:
        print("Main: Seeding transactions...", flush=True)
        db = SessionLocal()
        defaults = [
            {"nome_tela": "Dashboard", "rota": "/dashboard", "descricao": "Página inicial"},
            {"nome_tela": "Cadastros", "rota": "/cadastros", "descricao": "Área de cadastros"},
            {"nome_tela": "Empresas", "rota": "/companies", "descricao": "Gestão de empresas"},
            {"nome_tela": "Usuários", "rota": "/users", "descricao": "Gestão de usuários"},
            {"nome_tela": "Funcionários", "rota": "/employees", "descricao": "Gestão de funcionários"},
            {"nome_tela": "Unidades Gestoras", "rota": "/ugs", "descricao": "Gestão de UG"},
            {"nome_tela": "Centros de Custo", "rota": "/cost-centers", "descricao": "Gestão de centros de custo"},
            {"nome_tela": "Permissões", "rota": "/permissions", "descricao": "Controle de acessos"},
            {"nome_tela": "Revisões", "rota": "/reviews", "descricao": "Menu de revisões"},
            {"nome_tela": "Revisões - Períodos", "rota": "/reviews/periodos", "descricao": "Revisões por período"},
            {"nome_tela": "Revisões - Delegação", "rota": "/reviews/delegacao", "descricao": "Delegação de revisões"},
            {"nome_tela": "Revisões - Vidas Úteis", "rota": "/reviews/vidas-uteis", "descricao": "Revisão de vidas úteis"},
            {"nome_tela": "Revisões - Massa", "rota": "/revisoes-massa", "descricao": "Revisões em massa de ativos"},
            {"nome_tela": "Revisões - Cronogramas", "rota": "/reviews/cronogramas", "descricao": "Cronogramas de revisão"},
            {"nome_tela": "Revisões - Cronogramas (Edição)", "rota": "/reviews/cronogramas/edit", "descricao": "Permissão para editar cronogramas"},
            {"nome_tela": "Relatórios", "rota": "/reports", "descricao": "Relatórios"},
            {"nome_tela": "Ativos", "rota": "/assets", "descricao": "Gestão de ativos"},
        ]
        for item in defaults:
            exists = db.query(TransacaoModel).filter(TransacaoModel.rota == item["rota"]).first()
            if not exists:
                db.add(TransacaoModel(**item))
        db.commit()
        print("Main: Transactions seeded", flush=True)
    except Exception as e:
        # Seed errors shouldn't break startup; just log
        import traceback
        print("Seed transacoes error:", e)
        traceback.print_exc()
    finally:
        try:
            db.close()
        except Exception:
            pass

    # Opcional: seed do usuário administrador na primeira subida
    # Ative definindo a variável de ambiente AUTO_SEED_ADMIN=true
    try:
        auto_seed = os.getenv("AUTO_SEED_ADMIN", "").lower() in {"1", "true", "yes"}
        if auto_seed:
            # Import relativo ao pacote backend/scripts
            try:
                from ..scripts.seed_admin import ensure_admin
            except Exception:
                # Fallback: import absoluto quando o pacote raiz está no PYTHONPATH
                from scripts.seed_admin import ensure_admin  # type: ignore
            ensure_admin()
    except Exception as e:
        import traceback
        print("Seed admin error:", e)
        traceback.print_exc()

    try:
        print("Main: Ensuring admin user...", flush=True)
        db = SessionLocal()
        exists = db.query(UsuarioModel.id).limit(1).first()
        if not exists:
            pwd = os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123")
            senha_hash = bcrypt.hashpw(pwd.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            last = db.query(UsuarioModel.id).order_by(UsuarioModel.id.desc()).first()
            next_num = (last[0] if last else 0) + 1
            codigo = f"{next_num:06d}"
            u = UsuarioModel(
                codigo=codigo,
                nome_completo="Administrador",
                email=os.getenv("DEFAULT_ADMIN_EMAIL", "admin@local"),
                senha_hash=senha_hash,
                cpf=os.getenv("DEFAULT_ADMIN_CPF", "00000000000"),
                nome_usuario=os.getenv("DEFAULT_ADMIN_USERNAME", "admin"),
                status="Ativo",
            )
            db.add(u)
            db.commit()
    except Exception as e:
        import traceback
        print("Inline admin seed error:", e)
        traceback.print_exc()
    finally:
        try:
            db.close()
        except Exception:
            pass
    print("Main: on_startup completed", flush=True)

@app.get("/health")
def health():
    # Checagem real de conectividade com o banco
    try:
        with engine.connect() as conn:
            conn.execute(sa.text("SELECT 1"))
        return {"status": "ok", "db": "ok"}
    except UnicodeDecodeError:
        # Alguns clusters locais do PostgreSQL podem disparar erros de encoding
        # (ex.: LATIN1/WIN1252) ao conectar/consultar. Não é um erro de rede.
        # Para não travar o frontend com 503, retornamos OK com detalhe.
        return {"status": "ok", "db": "unknown", "detail": "encoding_issue"}
    except Exception as e:
        # Não expor credenciais, apenas o tipo do erro
        return JSONResponse(status_code=503, content={"status": "error", "db": "error", "detail": e.__class__.__name__})

@app.get("/debug/schema/{table_name}")
def debug_schema(table_name: str, db: Session = Depends(get_db)):
    try:
        insp = sa.inspect(engine)
        columns = [c["name"] for c in insp.get_columns(table_name)]
        return {"table": table_name, "columns": columns}
    except Exception as e:
        return {"error": str(e)}

@app.get("/")
def root():
    return {"message": "Asset Life API"}

# Endpoint /health já definido acima com verificação de banco de dados

# -----------------------------
# Util: Auditoria
# -----------------------------
def audit(db: Session, *, usuario_id: int | None, acao: str, entidade: str, entidade_id: int | None, detalhes: str | None = None):
    try:
        log = AuditoriaLogModel(
            usuario_id=usuario_id,
            acao=acao,
            entidade=entidade,
            entidade_id=entidade_id,
            detalhes=detalhes,
        )
        db.add(log)
        db.commit()
    except Exception:
        db.rollback()

# Schemas
class Company(BaseModel):
    id: int
    name: str
    cnpj: str
    branch_type: str
    street: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    cep: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    division: Optional[str] = None
    state_registration: Optional[str] = None
    status: str

    class Config:
        from_attributes = True  # pydantic v2 orm_mode

class CompanyCreate(BaseModel):
    name: str
    cnpj: str
    branch_type: str
    street: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    cep: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    division: Optional[str] = None
    state_registration: Optional[str] = None
    status: str

class CompanyUpdate(BaseModel):
    name: Optional[str] = None
    cnpj: Optional[str] = None
    branch_type: Optional[str] = None
    street: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    cep: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    division: Optional[str] = None
    state_registration: Optional[str] = None
    status: Optional[str] = None

# Companies CRUD (DB)
@app.get("/companies", response_model=List[Company])
def list_companies(current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    allowed = get_allowed_company_ids(db, current_user)
    if not allowed:
        return []
    return db.query(CompanyModel).filter(CompanyModel.id.in_(allowed)).all()

@app.get("/companies/{company_id}", response_model=Company)
def get_company(company_id: int, current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    c = db.query(CompanyModel).filter(CompanyModel.id == company_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")
    allowed = set(get_allowed_company_ids(db, current_user))
    if c.id not in allowed:
        raise HTTPException(status_code=403, detail="Acesso negado à empresa solicitada")
    return c

@app.post("/companies", response_model=Company)
def create_company(payload: CompanyCreate, db: Session = Depends(get_db)):
    c = CompanyModel(**payload.dict())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/companies/{company_id}", response_model=Company)
def update_company(company_id: int, payload: CompanyUpdate, db: Session = Depends(get_db)):
    c = db.query(CompanyModel).filter(CompanyModel.id == company_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")
    for k, v in payload.dict(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return c

@app.delete("/companies/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db)):
    c = db.query(CompanyModel).filter(CompanyModel.id == company_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")
    db.delete(c)
    db.commit()
    return {"deleted": True}

class Cronograma(BaseModel):
    id: int
    periodo_id: int
    empresa_id: int
    responsavel_id: int
    descricao: Optional[str] = None
    status: str
    progresso_percentual: int
    criado_em: datetime
    class Config:
        from_attributes = True

class CronogramaCreate(BaseModel):
    periodo_id: int
    empresa_id: int
    responsavel_id: int
    descricao: Optional[str] = None
    status: str = "Aberto"

class CronogramaUpdate(BaseModel):
    descricao: Optional[str] = None
    status: Optional[str] = None
    progresso_percentual: Optional[int] = None

class CronogramaTarefa(BaseModel):
    id: int
    cronograma_id: int
    tipo: str
    nome: str
    descricao: Optional[str] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    responsavel_id: Optional[int] = None
    status: str
    progresso_percentual: int
    dependente_tarefa_id: Optional[int] = None
    criado_em: Optional[datetime] = None
    class Config:
        from_attributes = True

class CronogramaTarefaCreate(BaseModel):
    tipo: Optional[str] = "Tarefa"
    nome: str
    descricao: Optional[str] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    responsavel_id: Optional[int] = None
    status: str = "Pendente"
    progresso_percentual: Optional[int] = 0
    dependente_tarefa_id: Optional[int] = None

class CronogramaTarefaUpdate(BaseModel):
    tipo: Optional[str] = None
    nome: Optional[str] = None
    descricao: Optional[str] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    responsavel_id: Optional[int] = None
    status: Optional[str] = None
    progresso_percentual: Optional[int] = None
    dependente_tarefa_id: Optional[int] = None

class CronogramaResumo(BaseModel):
    total_tarefas: int
    concluido: int
    em_andamento: int
    pendente: int
    atrasada: int
    progresso_percentual: int
    previsao_conclusao: Optional[date] = None

@app.get("/cronogramas", response_model=List[Cronograma])
def list_cronogramas(periodo_id: Optional[int] = None, empresa_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(CronogramaModel)
    if periodo_id is not None:
        q = q.filter(CronogramaModel.periodo_id == int(periodo_id))
    if empresa_id is not None:
        q = q.filter(CronogramaModel.empresa_id == int(empresa_id))
    return q.order_by(CronogramaModel.id.desc()).all()

@app.post("/cronogramas", response_model=Cronograma)
def create_cronograma(payload: CronogramaCreate, template: bool = True, db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    per = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == payload.periodo_id).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período não encontrado")
    if getattr(per, "status", None) == "Fechado":
        raise HTTPException(status_code=400, detail="Período fechado")
    existing = db.query(CronogramaModel).filter(CronogramaModel.periodo_id == payload.periodo_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Já existe um cronograma para este período")
    emp = db.query(CompanyModel).filter(CompanyModel.id == payload.empresa_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    resp = db.query(UsuarioModel).filter(UsuarioModel.id == payload.responsavel_id).first()
    if not resp:
        raise HTTPException(status_code=404, detail="Responsável não encontrado")
    c = CronogramaModel(
        periodo_id=payload.periodo_id,
        empresa_id=payload.empresa_id,
        responsavel_id=payload.responsavel_id,
        descricao=payload.descricao,
        status=payload.status,
    )
    try:
        db.add(c)
        db.commit()
        db.refresh(c)
    except sa.exc.IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Violação de integridade nos dados do cronograma")
    if template:
        inicio = getattr(per, "data_abertura", None)
        fim_prev = getattr(per, "data_fechamento_prevista", None)
        
        # 1. Tarefas estáticas antes da execução
        tasks_to_create = [
            {"tipo": "Título", "nome": "Iniciação", "desc": None},
            {"tipo": "Tarefa", "nome": "Reunião Kick Off", "desc": None},
            {"tipo": "Título", "nome": "Planejamento", "desc": None},
            {"tipo": "Tarefa", "nome": "Consolidação da Base de Dados", "desc": None},
            {"tipo": "Tarefa", "nome": "Abertura do Período", "desc": None},
            {"tipo": "Tarefa", "nome": "Delegação de itens", "desc": None},
            {"tipo": "Tarefa", "nome": "Treinamento", "desc": None},
            {"tipo": "Título", "nome": "Execução", "desc": None},
        ]

        # 2. Tarefas dinâmicas de Execução (Revisores)
        # Buscar delegações ativas neste período agrupadas por revisor
        try:
            delegations = db.query(RevisaoDelegacaoModel.revisor_id, sa.func.count(RevisaoDelegacaoModel.id))\
                .filter(RevisaoDelegacaoModel.periodo_id == payload.periodo_id, RevisaoDelegacaoModel.status == 'Ativo')\
                .group_by(RevisaoDelegacaoModel.revisor_id).all()
            
            for revisor_id, total_assigned in delegations:
                if total_assigned == 0: continue
                
                revisor = db.query(UsuarioModel).filter(UsuarioModel.id == revisor_id).first()
                if not revisor: continue
                
                # Calcular progresso (itens revisados / total)
                # Consideramos revisado se RevisaoItem.alterado == True
                reviewed_count = db.query(sa.func.count(RevisaoItemModel.id))\
                    .join(RevisaoDelegacaoModel, RevisaoDelegacaoModel.ativo_id == RevisaoItemModel.id)\
                    .filter(
                        RevisaoDelegacaoModel.periodo_id == payload.periodo_id,
                        RevisaoDelegacaoModel.revisor_id == revisor_id,
                        RevisaoDelegacaoModel.status == 'Ativo',
                        RevisaoItemModel.alterado == True
                    ).scalar() or 0
                    
                progresso = int((reviewed_count / total_assigned) * 100)
                
                tasks_to_create.append({
                    "tipo": "Tarefa",
                    "nome": f"Revisão de Vidas úteis - {revisor.nome_completo}",
                    "desc": f"Revisor: {revisor.nome_completo} ({reviewed_count}/{total_assigned} itens)",
                    "responsavel_id": revisor_id,
                    "progresso_percentual": progresso
                })
        except Exception as e:
            print(f"Erro ao gerar tarefas dinâmicas: {e}")
            # Fallback se falhar
            tasks_to_create.append({"tipo": "Tarefa", "nome": "Revisão Vidas Úteis", "desc": "Erro ao gerar lista de revisores"})

        # 3. Tarefas estáticas pós-execução
        tasks_to_create.extend([
            {"tipo": "Título", "nome": "Encerramento", "desc": None},
            {"tipo": "Tarefa", "nome": "Fechar Período", "desc": None},
            {"tipo": "Tarefa", "nome": "Confeccionar Laudo", "desc": None},
            {"tipo": "Tarefa", "nome": "Assinatura de Laudo", "desc": None},
            {"tipo": "Tarefa", "nome": "Abertura de Chamado", "desc": None},
            {"tipo": "Tarefa", "nome": "Ajustes no Sistema", "desc": None},
        ])

        has_tipo = False
        try:
            insp = sa.inspect(engine)
            cols = [c["name"] for c in insp.get_columns("cronogramas_tarefas")]
            has_tipo = "tipo" in cols
        except Exception:
            has_tipo = False
            
        try:
            for task in tasks_to_create:
                tipo = task.get("tipo", "Tarefa")
                nome = task.get("nome")
                desc = task.get("desc")
                resp_id = task.get("responsavel_id") # Pode ser None
                prog = task.get("progresso_percentual", 0)

                if has_tipo:
                    t = CronogramaTarefaModel(
                        cronograma_id=c.id,
                        tipo=tipo,
                        nome=nome,
                        descricao=desc,
                        data_inicio=inicio,
                        data_fim=fim_prev,
                        responsavel_id=resp_id,
                        status="Pendente",
                        progresso_percentual=prog,
                    )
                    db.add(t)
                else:
                    # Fallback for when 'tipo' column is missing: prepend Título to name if needed
                    final_nome = f"[TÍTULO] {nome}" if tipo == "Título" else nome
                    db.execute(sa.text(
                        """
                        INSERT INTO cronogramas_tarefas (
                            cronograma_id, nome, descricao, data_inicio, data_fim,
                            responsavel_id, status, progresso_percentual, dependente_tarefa_id
                        ) VALUES (
                            :cronograma_id, :nome, :descricao, :data_inicio, :data_fim,
                            :responsavel_id, :status, :progresso_percentual, :dependente_tarefa_id
                        )
                        """
                    ), {
                        "cronograma_id": c.id,
                        "nome": final_nome,
                        "descricao": desc,
                        "data_inicio": inicio,
                        "data_fim": fim_prev,
                        "responsavel_id": resp_id,
                        "status": "Pendente",
                        "progresso_percentual": prog,
                        "dependente_tarefa_id": None,
                    })
            db.commit()
        except Exception:
            db.rollback()
            raise HTTPException(status_code=400, detail="Falha ao criar tarefas padrão do cronograma")
    return c

@app.get("/cronogramas/{cronograma_id}", response_model=Cronograma)
def get_cronograma(cronograma_id: int, db: Session = Depends(get_db)):
    c = db.query(CronogramaModel).filter(CronogramaModel.id == cronograma_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cronograma não encontrado")
    return c

@app.put("/cronogramas/{cronograma_id}", response_model=Cronograma)
def update_cronograma(cronograma_id: int, payload: CronogramaUpdate, db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    c = db.query(CronogramaModel).filter(CronogramaModel.id == cronograma_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cronograma não encontrado")
    
    if payload.descricao is not None:
        c.descricao = payload.descricao
    
    if payload.status is not None:
        # Check tasks if closing
        if payload.status == 'Concluído':
            pending = db.execute(sa.text(
                "SELECT COUNT(*) FROM cronogramas_tarefas WHERE cronograma_id = :cid AND status != 'Concluída'"
            ), {"cid": cronograma_id}).scalar()
            if pending > 0:
                raise HTTPException(status_code=400, detail="Não é possível concluir o cronograma com tarefas pendentes.")
        c.status = payload.status

    if payload.progresso_percentual is not None:
        c.progresso_percentual = payload.progresso_percentual
        
    db.commit()
    db.refresh(c)
    return c

@app.get("/cronogramas/{cronograma_id}/tarefas", response_model=List[CronogramaTarefa])
def list_cronograma_tarefas(cronograma_id: int, db: Session = Depends(get_db)):
    log_path = os.path.join(os.path.dirname(__file__), "..", "debug_log.txt")
    def log(msg):
        try:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now()}] {msg}\n")
        except: pass

    log(f"list_cronograma_tarefas called for ID: {cronograma_id}")

    try:
        # Check if cronograma exists
        c = db.query(CronogramaModel).filter(CronogramaModel.id == cronograma_id).first()
        if not c:
            raise HTTPException(status_code=404, detail="Cronograma não encontrado")

        # Calculate progress from revisions (Fix for Julio Cesar Carvalho issue)
        stats_map = {}
        try:
            if c.periodo_id:
                stats_query = sa.text("""
                    SELECT d.revisor_id, 
                           COUNT(*) as total, 
                           SUM(CASE WHEN (i.status IN ('Revisado', 'Aprovado') OR i.alterado = TRUE) THEN 1 ELSE 0 END) as revisados
                    FROM revisoes_delegacoes d
                    JOIN revisoes_itens i ON d.ativo_id = i.id
                    WHERE d.periodo_id = :pid
                    GROUP BY d.revisor_id
                """)
                stats_rows = db.execute(stats_query, {"pid": c.periodo_id}).mappings().all()
                stats_map = {row['revisor_id']: row for row in stats_rows}
                log(f"Stats map calculated: {len(stats_map)} entries")

                # SYNC: Check for missing revisors in the schedule
                # Find all revisors with active delegations in this period
                active_revisors = db.query(RevisaoDelegacaoModel.revisor_id).filter(
                    RevisaoDelegacaoModel.periodo_id == c.periodo_id,
                    RevisaoDelegacaoModel.status == 'Ativo'
                ).distinct().all()
                active_revisors_ids = set(r[0] for r in active_revisors)

                # Find revisors who already have a task in this cronograma
                existing_task_revisors = db.query(CronogramaTarefaModel.responsavel_id).filter(
                    CronogramaTarefaModel.cronograma_id == cronograma_id,
                    CronogramaTarefaModel.nome.like('Revisão de Vidas úteis - %')
                ).all()
                existing_revisors_ids = set(r[0] for r in existing_task_revisors if r[0] is not None)

                # Identify missing revisors
                missing_ids = active_revisors_ids - existing_revisors_ids
                
                if missing_ids:
                    log(f"Found {len(missing_ids)} missing revisors in schedule. Syncing...")
                    for rid in missing_ids:
                        user = db.query(UsuarioModel).filter(UsuarioModel.id == rid).first()
                        if not user: continue
                        
                        # Calculate initial progress
                        s = stats_map.get(rid, {'total': 0, 'revisados': 0})
                        progresso = 0
                        if s['total'] > 0:
                            progresso = int(round((s['revisados'] / s['total']) * 100))
                        
                        # Get period dates for task duration
                        per = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == c.periodo_id).first()
                        data_inicio = getattr(per, "data_abertura", None)
                        data_fim = getattr(per, "data_fechamento_prevista", None)

                        try:
                            new_task = CronogramaTarefaModel(
                                cronograma_id=cronograma_id,
                                tipo="Tarefa",
                                nome=f"Revisão de Vidas úteis - {user.nome_completo}",
                                descricao=f"Revisor: {user.nome_completo}",
                                data_inicio=data_inicio,
                                data_fim=data_fim,
                                responsavel_id=rid,
                                status="Concluída" if progresso == 100 else ("Em Andamento" if progresso > 0 else "Pendente"),
                                progresso_percentual=progresso
                            )
                            db.add(new_task)
                            db.commit()
                        except (sa.exc.ProgrammingError, sa.exc.DBAPIError, sa.exc.IntegrityError):
                            db.rollback()
                            # Fallback if 'tipo' column is missing or other DB issue
                            log(f"Fallback insert for revisor {rid}")
                            db.execute(sa.text(
                                """
                                INSERT INTO cronogramas_tarefas (
                                    cronograma_id, nome, descricao, data_inicio, data_fim,
                                    responsavel_id, status, progresso_percentual
                                ) VALUES (
                                    :cid, :nome, :desc, :di, :df, :rid, :st, :prog
                                )
                                """
                            ), {
                                "cid": cronograma_id,
                                "nome": f"Revisão de Vidas úteis - {user.nome_completo}",
                                "desc": f"Revisor: {user.nome_completo}",
                                "di": data_inicio,
                                "df": data_fim,
                                "rid": rid,
                                "st": "Concluída" if progresso == 100 else ("Em Andamento" if progresso > 0 else "Pendente"),
                                "prog": progresso
                            })
                            db.commit()
                    
                    log("Sync completed.")

        except Exception as e_stats:
            log(f"Error calculating stats or syncing revisors: {e_stats}")

        # Explicitly define columns that are known to exist in the migration
        # We exclude 'tipo' because it is not in the migration.
        db_cols = [
            "id", "cronograma_id", "nome", "descricao",
            "data_inicio", "data_fim", "responsavel_id", 
            "status", "progresso_percentual", "dependente_tarefa_id",
            "criado_em"
        ]
        
        cols_str = ", ".join(db_cols)
        sql = sa.text(f"SELECT {cols_str} FROM cronogramas_tarefas WHERE cronograma_id = :cid ORDER BY id")

        rows = []
        try:
            # Try full fetch first
            rows = db.execute(sql, {"cid": cronograma_id}).mappings().all()
            log(f"Full fetch success, rows: {len(rows)}")
        except Exception as e:
            log(f"Full fetch failed: {e}")
            print(f"Full fetch failed: {e}")
            # Fallback: Minimal fetch (id, nome, status, data_fim) which are critical
            try:
                sql_min = sa.text("SELECT id, cronograma_id, nome, status, data_fim FROM cronogramas_tarefas WHERE cronograma_id = :cid")
                rows = db.execute(sql_min, {"cid": cronograma_id}).mappings().all()
                log(f"Minimal fetch success, rows: {len(rows)}")
            except Exception as e2:
                log(f"Minimal fetch failed: {e2}")
                print(f"Minimal fetch failed: {e2}")
                # Try absolute minimal fetch to verify connectivity/existence
                try:
                    count = db.execute(sa.text("SELECT count(*) FROM cronogramas_tarefas WHERE cronograma_id = :cid"), {"cid": cronograma_id}).scalar()
                    log(f"Count check: {count}")
                    if count > 0:
                        raise HTTPException(status_code=500, detail=f"Erro crítico: Existem {count} tarefas, mas falha ao listar colunas. Erro: {str(e2)}")
                except Exception as e3:
                    log(f"Count check failed: {e3}")
                
                raise HTTPException(status_code=500, detail=f"Erro ao buscar tarefas (min): {str(e2)}")

        now = date.today()
        result = []
        last_error = None
        for r in rows:
            try:
                # Safe extraction of values
                status = r.get("status") or "Pendente"
                df = r.get("data_fim")
                rid = r.get("responsavel_id")

                # Dynamic progress calculation (Only for revision tasks)
                progresso = r.get("progresso_percentual") or 0
                # Only auto-calculate if it's a revision task
                is_revision_task = r.get("nome", "").startswith("Revisão de Vidas úteis")
                
                if is_revision_task and rid and rid in stats_map:
                    s = stats_map[rid]
                    if s['total'] > 0:
                        progresso = int(round((s['revisados'] / s['total']) * 100))
                        # Update DB if different
                        if progresso != (r.get("progresso_percentual") or 0):
                            try:
                                db.execute(sa.text("UPDATE cronogramas_tarefas SET progresso_percentual = :p WHERE id = :id"), {"p": progresso, "id": r["id"]})
                                db.commit()
                            except Exception as e_upd:
                                log(f"Failed to update progress for task {r['id']}: {e_upd}")
                
                # Auto-complete status if 100% (prioritize over delay)
                # Only for revision tasks to avoid overriding manual status of other tasks
                if is_revision_task:
                    if progresso == 100:
                        status = "Concluída"
                        # Persist completed status if not already
                        if r.get("status") != "Concluída":
                            try:
                                db.execute(sa.text("UPDATE cronogramas_tarefas SET status = 'Concluída' WHERE id = :id"), {"id": r["id"]})
                                db.commit()
                            except Exception as e_st:
                                log(f"Failed to update status to Concluída for task {r['id']}: {e_st}")
                    elif progresso < 100 and status == "Concluída":
                        status = "Em Andamento"
                        try:
                            db.execute(sa.text("UPDATE cronogramas_tarefas SET status = 'Em Andamento' WHERE id = :id"), {"id": r["id"]})
                            db.commit()
                        except Exception: pass

                # Check for delay (only if not completed)
                if df and isinstance(df, date) and df < now and status != "Concluída":
                    status = "Atrasada"

                # Construct Pydantic model safely
                task = CronogramaTarefa(
                    id=r["id"], 
                    cronograma_id=r["cronograma_id"], 
                    tipo="Tarefa", # Hardcoded as it's missing in DB
                    nome=r["nome"],
                    descricao=r.get("descricao"), 
                    data_inicio=r.get("data_inicio"), 
                    data_fim=df,
                    responsavel_id=rid, 
                    status=status,
                    progresso_percentual=progresso,
                    dependente_tarefa_id=r.get("dependente_tarefa_id"),
                    criado_em=r.get("criado_em")
                )
                result.append(task)
            except Exception as loop_e:
                last_error = loop_e
                log(f"Error creating task object: {loop_e}")
                continue

        if not result and len(rows) > 0:
            # If we have rows but no result, it means all failed.
            keys = list(rows[0].keys()) if rows else "No rows"
            log(f"All rows failed conversion. Keys: {keys}. Last error: {last_error}")
            raise HTTPException(status_code=500, detail=f"Erro de validação dos dados (chaves: {keys}): {str(last_error)}")

        return result
    except HTTPException:
        raise
    except Exception as e:
        log(f"Critical error in list_cronograma_tarefas: {e}")
        print(f"Critical error in list_cronograma_tarefas: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno ao listar tarefas: {str(e)}")

@app.post("/cronogramas/{cronograma_id}/tarefas", response_model=CronogramaTarefa)
def create_cronograma_tarefa(cronograma_id: int, payload: CronogramaTarefaCreate, db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    c = db.query(CronogramaModel).filter(CronogramaModel.id == cronograma_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cronograma não encontrado")
    try:
        t = CronogramaTarefaModel(
            cronograma_id=cronograma_id,
            tipo=payload.tipo or "Tarefa",
            nome=payload.nome,
            descricao=payload.descricao,
            data_inicio=payload.data_inicio,
            data_fim=payload.data_fim,
            responsavel_id=payload.responsavel_id,
            status=payload.status,
            progresso_percentual=payload.progresso_percentual or 0,
            dependente_tarefa_id=payload.dependente_tarefa_id,
        )
        db.add(t)
        db.commit()
        db.refresh(t)
        return t
    except (sa.exc.ProgrammingError, sa.exc.DBAPIError):
        db.rollback()
        # Fallback quando coluna 'tipo' não existe
        try:
            db.execute(sa.text(
                """
                INSERT INTO cronogramas_tarefas (
                    cronograma_id, nome, descricao, data_inicio, data_fim,
                    responsavel_id, status, progresso_percentual, dependente_tarefa_id
                ) VALUES (
                    :cronograma_id, :nome, :descricao, :data_inicio, :data_fim,
                    :responsavel_id, :status, :progresso_percentual, :dependente_tarefa_id
                )
                RETURNING id
                """
            ), {
                "cronograma_id": cronograma_id,
                "nome": payload.nome,
                "descricao": payload.descricao,
                "data_inicio": payload.data_inicio,
                "data_fim": payload.data_fim,
                "responsavel_id": payload.responsavel_id,
                "status": payload.status,
                "progresso_percentual": payload.progresso_percentual or 0,
                "dependente_tarefa_id": payload.dependente_tarefa_id,
            })
            db.commit()
            # Recupera a tarefa criada para retorno coerente
            created = db.execute(sa.text("SELECT * FROM cronogramas_tarefas WHERE cronograma_id = :cid ORDER BY id DESC LIMIT 1"), {"cid": cronograma_id}).mappings().first()
            if not created:
                raise HTTPException(status_code=400, detail="Falha ao criar tarefa")
            # Monta resposta mínima, com 'tipo' como 'Tarefa' por compatibilidade
            return CronogramaTarefa(
                id=created["id"], cronograma_id=created["cronograma_id"], tipo="Tarefa", nome=created["nome"],
                descricao=created.get("descricao"), data_inicio=created.get("data_inicio"), data_fim=created.get("data_fim"),
                responsavel_id=created.get("responsavel_id"), status=created.get("status"),
                progresso_percentual=created.get("progresso_percentual") or 0,
                dependente_tarefa_id=created.get("dependente_tarefa_id"), criado_em=created.get("criado_em")
            )
        except Exception:
            raise HTTPException(status_code=400, detail="Falha ao criar tarefa")

@app.put("/cronogramas/{cronograma_id}/tarefas/{tarefa_id}", response_model=CronogramaTarefa)
def update_cronograma_tarefa(cronograma_id: int, tarefa_id: int, payload: CronogramaTarefaUpdate, db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    try:
        t = db.query(CronogramaTarefaModel).filter(CronogramaTarefaModel.id == tarefa_id, CronogramaTarefaModel.cronograma_id == cronograma_id).first()
        if not t:
            raise HTTPException(status_code=404, detail="Tarefa não encontrada")
        data = payload.dict(exclude_unset=True)
        for k, v in data.items():
            setattr(t, k, v)
        db.commit()
        db.refresh(t)
        return t
    except (sa.exc.ProgrammingError, sa.exc.DBAPIError) as e:
        log(f"Error updating task {tarefa_id}: {e}")
        db.rollback()
        # Fallback raw SQL update due to missing columns (e.g. tipo)
        data = payload.dict(exclude_unset=True)
        params = {"id": tarefa_id, "cid": cronograma_id}
        clauses = []
        for k, v in data.items():
            clauses.append(f"{k} = :{k}")
            params[k] = v
        
        if clauses:
            sql = sa.text(f"UPDATE cronogramas_tarefas SET {', '.join(clauses)} WHERE id = :id AND cronograma_id = :cid")
            res = db.execute(sql, params)
            db.commit()
            if res.rowcount == 0:
                raise HTTPException(status_code=404, detail="Tarefa não encontrada")
        
        # Fetch updated
        cols = "id, cronograma_id, tipo, nome, descricao, data_inicio, data_fim, responsavel_id, status, progresso_percentual, dependente_tarefa_id, criado_em"
        row = db.execute(sa.text(f"SELECT {cols} FROM cronogramas_tarefas WHERE id=:id"), {"id": tarefa_id}).mappings().first()
        if not row:
             raise HTTPException(status_code=404, detail="Tarefa não encontrada")
        return CronogramaTarefa(tipo=row.get("tipo") or "Tarefa", **row)

# Evidências de tarefas
class CronogramaTarefaEvidencia(BaseModel):
    id: int
    tarefa_id: int
    nome_arquivo: str
    content_type: str
    tamanho_bytes: int
    criado_em: datetime
    uploaded_by: Optional[int] = None
    class Config:
        from_attributes = True

@app.get("/cronogramas/{cronograma_id}/tarefas/{tarefa_id}/evidencias", response_model=List[CronogramaTarefaEvidencia])
def list_evidencias(cronograma_id: int, tarefa_id: int, db: Session = Depends(get_db)):
    try:
        # Use Raw SQL to check task existence to avoid missing column errors
        exists = db.execute(sa.text("SELECT id FROM cronogramas_tarefas WHERE id=:id AND cronograma_id=:cid"), {"id": tarefa_id, "cid": cronograma_id}).scalar()
        if not exists:
             raise HTTPException(status_code=404, detail="Tarefa não encontrada")
             
        items = db.query(CronogramaTarefaEvidenciaModel).filter(CronogramaTarefaEvidenciaModel.tarefa_id == tarefa_id).order_by(CronogramaTarefaEvidenciaModel.id.desc()).all()
        return items
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error listing evidencias: {e}")
        return []

@app.post("/cronogramas/{cronograma_id}/tarefas/{tarefa_id}/evidencias", response_model=CronogramaTarefaEvidencia)
async def upload_evidencia(cronograma_id: int, tarefa_id: int, file: UploadFile = File(...), current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    # Raw SQL check
    exists = db.execute(sa.text("SELECT id FROM cronogramas_tarefas WHERE id=:id AND cronograma_id=:cid"), {"id": tarefa_id, "cid": cronograma_id}).scalar()
    if not exists:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio")
    item = CronogramaTarefaEvidenciaModel(
        tarefa_id=tarefa_id,
        nome_arquivo=file.filename or "arquivo",
        content_type=file.content_type or "application/octet-stream",
        tamanho_bytes=len(content),
        conteudo=content,
        uploaded_by=getattr(current_user, "id", None),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

@app.get("/cronogramas/{cronograma_id}/tarefas/{tarefa_id}/evidencias/{evidencia_id}")
def download_evidencia(cronograma_id: int, tarefa_id: int, evidencia_id: int, db: Session = Depends(get_db)):
    exists = db.execute(sa.text("SELECT id FROM cronogramas_tarefas WHERE id=:id AND cronograma_id=:cid"), {"id": tarefa_id, "cid": cronograma_id}).scalar()
    if not exists:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada")
    ev = db.query(CronogramaTarefaEvidenciaModel).filter(CronogramaTarefaEvidenciaModel.id == evidencia_id, CronogramaTarefaEvidenciaModel.tarefa_id == tarefa_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Evidência não encontrada")
    headers = {
        "Content-Disposition": f"attachment; filename=\"{ev.nome_arquivo}\""
    }
    return Response(content=ev.conteudo, media_type=ev.content_type or "application/octet-stream", headers=headers)

@app.delete("/cronogramas/{cronograma_id}/tarefas/{tarefa_id}/evidencias/{evidencia_id}")
def delete_evidencia(cronograma_id: int, tarefa_id: int, evidencia_id: int, db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    check_permission(db, current_user, "/reviews/cronogramas/edit")
    exists = db.execute(sa.text("SELECT id FROM cronogramas_tarefas WHERE id=:id AND cronograma_id=:cid"), {"id": tarefa_id, "cid": cronograma_id}).scalar()
    if not exists:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada")
    ev = db.query(CronogramaTarefaEvidenciaModel).filter(CronogramaTarefaEvidenciaModel.id == evidencia_id, CronogramaTarefaEvidenciaModel.tarefa_id == tarefa_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Evidência não encontrada")
    db.delete(ev)
    db.commit()
    return {"deleted": True}

@app.get("/cronogramas/{cronograma_id}/resumo", response_model=CronogramaResumo)
def cronograma_resumo(cronograma_id: int, db: Session = Depends(get_db)):
    try:
        c = db.query(CronogramaModel).filter(CronogramaModel.id == cronograma_id).first()
        if not c:
            raise HTTPException(status_code=404, detail="Cronograma não encontrado")

        rows = db.execute(sa.text(
            "SELECT status, data_fim, progresso_percentual FROM cronogramas_tarefas WHERE cronograma_id = :cid"
        ), {"cid": cronograma_id}).mappings().all()

        now = date.today()
        total = len(rows)
        pendente = em_andamento = concluido = atrasada = 0
        progresso_media = 0
        for r in rows:
            st = r["status"] or "Pendente"
            if st == "Pendente":
                pendente += 1
            elif st == "Em Andamento":
                em_andamento += 1
            elif st == "Concluída":
                concluido += 1
            else:
                pendente += 1
            df = r["data_fim"]
            if df and isinstance(df, date) and df < now and st != "Concluída":
                atrasada += 1
            progresso_media += (r.get("progresso_percentual") or 0)
        progresso_percentual = int(round(progresso_media / total)) if total else 0
        return CronogramaResumo(
            total_tarefas=total,
            concluido=concluido,
            em_andamento=em_andamento,
            pendente=pendente,
            atrasada=atrasada,
            progresso_percentual=progresso_percentual,
            previsao_conclusao=None,
        )
    except HTTPException:
        raise
    except Exception:
        return CronogramaResumo(
            total_tarefas=0,
            concluido=0,
            em_andamento=0,
            pendente=0,
            atrasada=0,
            progresso_percentual=0,
            previsao_conclusao=None,
        )
    total = len(items)
    concl = sum(1 for i in items if i.status == "Concluída")
    em = sum(1 for i in items if i.status == "Em Andamento")
    pend = sum(1 for i in items if i.status == "Pendente")
    atr = sum(1 for i in items if i.status == "Atrasada")
    prog = int(round((concl / total) * 100)) if total > 0 else 0
    previsao = None
    if items:
        fins = [i.data_fim for i in items if i.data_fim is not None]
        previsao = max(fins) if fins else None
    return CronogramaResumo(
        total_tarefas=total,
        concluido=concl,
        em_andamento=em,
        pendente=pend,
        atrasada=atr,
        progresso_percentual=prog,
        previsao_conclusao=previsao,
    )
# Schemas de Colaboradores
class Colaborador(BaseModel):
    id: int
    full_name: str
    cpf: str
    matricula: Optional[str] = None
    cargo_funcao: Optional[str] = None
    empresa_id: int
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    tipo_vinculo: VinculoEnum
    data_admissao: date
    data_desligamento: Optional[date] = None
    telefone: Optional[str] = None
    email_corporativo: Optional[EmailStr] = None
    endereco: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    status: StatusEnum
    observacoes: Optional[str] = None

    class Config:
        from_attributes = True

class ColaboradorCreate(BaseModel):
    full_name: str
    cpf: str
    matricula: Optional[str] = None
    cargo_funcao: Optional[str] = None
    empresa_id: int
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    tipo_vinculo: VinculoEnum
    data_admissao: date
    data_desligamento: Optional[date] = None
    telefone: Optional[str] = None
    email_corporativo: Optional[EmailStr] = None
    endereco: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    status: StatusEnum = StatusEnum.ativo
    observacoes: Optional[str] = None

class ColaboradorUpdate(BaseModel):
    full_name: Optional[str] = None
    cpf: Optional[str] = None
    matricula: Optional[str] = None
    cargo_funcao: Optional[str] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    tipo_vinculo: Optional[VinculoEnum] = None
    data_admissao: Optional[date] = None
    data_desligamento: Optional[date] = None
    telefone: Optional[str] = None
    email_corporativo: Optional[EmailStr] = None
    endereco: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    status: Optional[StatusEnum] = None
    observacoes: Optional[str] = None

# CRUD de Colaboradores
@app.get("/colaboradores", response_model=List[Colaborador])
def list_colaboradores(request: Request, db: Session = Depends(get_db)):
    empresa_hdr = request.headers.get("X-Company-Id")
    q = db.query(EmployeeModel)
    if empresa_hdr:
        try:
            empresa_id = int(empresa_hdr)
            q = q.filter(EmployeeModel.empresa_id == empresa_id)
        except ValueError:
            pass
    return q.all()

@app.get("/colaboradores/{colab_id}", response_model=Colaborador)
def get_colaborador(colab_id: int, db: Session = Depends(get_db)):
    c = db.query(EmployeeModel).filter(EmployeeModel.id == colab_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    return c

@app.post("/colaboradores", response_model=Colaborador)
def create_colaborador(payload: ColaboradorCreate, db: Session = Depends(get_db)):
    c = EmployeeModel(**payload.dict())
    # validar empresa existente
    if not db.query(CompanyModel).filter(CompanyModel.id == c.empresa_id).first():
        raise HTTPException(status_code=400, detail="Empresa inválida")
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/colaboradores/{colab_id}", response_model=Colaborador)
def update_colaborador(colab_id: int, payload: ColaboradorUpdate, db: Session = Depends(get_db)):
    c = db.query(EmployeeModel).filter(EmployeeModel.id == colab_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    # se empresa_id foi enviado, valida
    data = payload.dict(exclude_unset=True)
    if "empresa_id" in data and data["empresa_id"] is not None:
        if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
            raise HTTPException(status_code=400, detail="Empresa inválida")
    for k, v in data.items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return c

@app.delete("/colaboradores/{colab_id}")
def delete_colaborador(colab_id: int, db: Session = Depends(get_db)):
    c = db.query(EmployeeModel).filter(EmployeeModel.id == colab_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    db.delete(c)
    db.commit()
    return {"deleted": True}

# -----------------------------
# Unidades Gerenciais (UG) - Schemas e CRUD
# -----------------------------
class UG(BaseModel):
    id: int
    codigo: str
    nome: str
    tipo_unidade: str
    nivel_hierarquico: str
    empresa_id: int
    responsavel_id: Optional[int] = None
    ug_superior_id: Optional[int] = None
    status: str

    class Config:
        from_attributes = True

class UGCreate(BaseModel):
    nome: str
    tipo_unidade: str
    nivel_hierarquico: str
    empresa_id: int
    responsavel_id: Optional[int] = None
    ug_superior_id: Optional[int] = None
    status: str = "Ativo"

class UGUpdate(BaseModel):
    nome: Optional[str] = None
    tipo_unidade: Optional[str] = None
    nivel_hierarquico: Optional[str] = None
    empresa_id: Optional[int] = None
    responsavel_id: Optional[int] = None
    ug_superior_id: Optional[int] = None
    status: Optional[str] = None

# Geração automática do código hierárquico
from sqlalchemy import func

def generate_codigo(db: Session, parent_id: Optional[int]) -> str:
    # raiz (sem pai): usa numeração sequencial no nível raiz
    if parent_id is None:
        count = db.query(func.count(UGModel.id)).filter(UGModel.ug_superior_id == None).scalar() or 0
        return f"{count + 1}"
    # com pai: prefixo é o código do pai e sufixo é sequencial dos filhos
    parent = db.query(UGModel).filter(UGModel.id == parent_id).first()
    if not parent:
        raise HTTPException(status_code=400, detail="UG superior inválida")
    siblings_count = db.query(func.count(UGModel.id)).filter(UGModel.ug_superior_id == parent_id).scalar() or 0
    return f"{parent.codigo}.{siblings_count + 1}"

# Validações de regras de negócio
VALID_LEVELS = {"CEO", "Diretoria", "Gerência Geral", "Gerência", "Coordenação", "Operacional"}
VALID_TYPES = {"Administrativa", "Produtiva", "Apoio", "Auxiliares"}

@app.get("/unidades_gerenciais", response_model=List[UG])
def list_ugs(request: Request, db: Session = Depends(get_db)):
    empresa_hdr = request.headers.get("X-Company-Id")
    q = db.query(UGModel)
    if empresa_hdr:
        try:
            empresa_id = int(empresa_hdr)
            q = q.filter(UGModel.empresa_id == empresa_id)
        except ValueError:
            pass
    return q.order_by(UGModel.codigo).all()

@app.get("/unidades_gerenciais/{ug_id}", response_model=UG)
def get_ug(ug_id: int, db: Session = Depends(get_db)):
    c = db.query(UGModel).filter(UGModel.id == ug_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="UG não encontrada")
    return c

@app.post("/unidades_gerenciais", response_model=UG)
def create_ug(payload: UGCreate, db: Session = Depends(get_db)):
    data = payload.dict()

    # validações de negócio
    if not data.get("nome"):
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    if data.get("nivel_hierarquico") not in VALID_LEVELS:
        raise HTTPException(status_code=400, detail="Nível hierárquico inválido")
    if data.get("tipo_unidade") not in VALID_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de unidade inválido")

    # CEO não pode ter UG superior
    if data["nivel_hierarquico"] == "CEO" and data.get("ug_superior_id"):
        raise HTTPException(status_code=400, detail="UG superior não é permitido para nível CEO")
    # outros níveis devem ter pai
    if data["nivel_hierarquico"] != "CEO" and not data.get("ug_superior_id"):
        raise HTTPException(status_code=400, detail="UG superior é obrigatório exceto para CEO")

    # valida empresa existente
    if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
        raise HTTPException(status_code=400, detail="Empresa inválida")
    # valida responsável se enviado
    if data.get("responsavel_id") and not db.query(EmployeeModel).filter(EmployeeModel.id == data["responsavel_id"]).first():
        raise HTTPException(status_code=400, detail="Responsável inválido")

    codigo = generate_codigo(db, data.get("ug_superior_id"))
    c = UGModel(codigo=codigo, **data)
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/unidades_gerenciais/{ug_id}", response_model=UG)
def update_ug(ug_id: int, payload: UGUpdate, db: Session = Depends(get_db)):
    c = db.query(UGModel).filter(UGModel.id == ug_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="UG não encontrada")

    data = payload.dict(exclude_unset=True)

    # validações básicas se enviados
    if "nivel_hierarquico" in data and data["nivel_hierarquico"] not in VALID_LEVELS:
        raise HTTPException(status_code=400, detail="Nível hierárquico inválido")
    if "tipo_unidade" in data and data["tipo_unidade"] not in VALID_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de unidade inválido")

    # regra de UG superior vs nível
    if data.get("nivel_hierarquico") == "CEO" and (data.get("ug_superior_id") or c.ug_superior_id):
        raise HTTPException(status_code=400, detail="UG superior não é permitido para nível CEO")

    if "empresa_id" in data and data["empresa_id"] is not None:
        if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
            raise HTTPException(status_code=400, detail="Empresa inválida")

    if "responsavel_id" in data and data["responsavel_id"] is not None:
        if not db.query(EmployeeModel).filter(EmployeeModel.id == data["responsavel_id"]).first():
            raise HTTPException(status_code=400, detail="Responsável inválido")

    # não recalculamos código automaticamente em update para evitar quebra
    for k, v in data.items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return c

@app.delete("/unidades_gerenciais/{ug_id}")
def delete_ug(ug_id: int, db: Session = Depends(get_db)):
    c = db.query(UGModel).filter(UGModel.id == ug_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="UG não encontrada")
    db.delete(c)
    db.commit()
    return {"deleted": True}

# -----------------------------
# Centros de Custos - Schemas e Endpoints
# -----------------------------
class CentroCusto(BaseModel):
    id: int
    codigo: str
    nome: str
    empresa_id: int
    ug_id: int
    responsavel_id: Optional[int] = None
    observacoes: Optional[str] = None
    data_criacao: Optional[datetime] = None
    criado_por: Optional[int] = None
    status: str
    class Config:
        from_attributes = True

class CentroCustoCreate(BaseModel):
    nome: str
    empresa_id: int
    ug_id: int
    responsavel_id: Optional[int] = None
    observacoes: Optional[str] = None
    criado_por: Optional[int] = None
    status: str = "Ativo"

@app.get("/centros_custos", response_model=List[CentroCusto])
def list_centros_custos(
    empresa_id: Optional[int] = None,
    ug_id: Optional[int] = None,
    nome: Optional[str] = None,
    status: Optional[str] = None,
    current_user: UsuarioModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    q = db.query(CentroCustoModel)
    # limitar às empresas permitidas do usuário
    allowed = get_allowed_company_ids(db, current_user)
    if allowed:
        q = q.filter(CentroCustoModel.empresa_id.in_(allowed))
    if empresa_id:
        # bloquear acesso explícito a empresa não permitida
        if allowed and empresa_id not in allowed:
            raise HTTPException(status_code=403, detail="Acesso negado aos Centros de Custos da empresa informada")
        q = q.filter(CentroCustoModel.empresa_id == empresa_id)
    if ug_id:
        q = q.filter(CentroCustoModel.ug_id == ug_id)
    if status:
        q = q.filter(CentroCustoModel.status == status)
    if nome:
        q = q.filter(sa.func.lower(CentroCustoModel.nome).like(f"%{nome.lower()}%"))
    return q.order_by(CentroCustoModel.codigo).all()

@app.get("/centros_custos/{cc_id}", response_model=CentroCusto)
def get_centro_custo(cc_id: int, db: Session = Depends(get_db)):
    c = db.query(CentroCustoModel).filter(CentroCustoModel.id == cc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Centro de Custos não encontrado")
    return c

@app.post("/centros_custos", response_model=CentroCusto)
def create_centro_custo(payload: CentroCustoCreate, db: Session = Depends(get_db)):
    data = payload.dict()

    # valida empresa e UG
    empresa = db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first()
    if not empresa:
        raise HTTPException(status_code=400, detail="Empresa inválida")

    ug = db.query(UGModel).filter(UGModel.id == data["ug_id"]).first()
    if not ug:
        raise HTTPException(status_code=400, detail="UG inválida")

    if ug.empresa_id != empresa.id:
        raise HTTPException(status_code=400, detail="UG não pertence à empresa informada")

    # valida responsável se informado
    if data.get("responsavel_id"):
        if not db.query(EmployeeModel).filter(EmployeeModel.id == data["responsavel_id"]).first():
            raise HTTPException(status_code=400, detail="Responsável inválido")

    # evitar duplicidade de nome dentro da mesma UG
    existing = (
        db.query(CentroCustoModel)
        .filter(sa.func.lower(CentroCustoModel.nome) == data["nome"].lower())
        .filter(CentroCustoModel.ug_id == data["ug_id"])
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Já existe um Centro de Custos com este nome na UG")

    # geração automática de código: <codigo_ug>.<sequencial>
    ultimo = (
        db.query(CentroCustoModel)
        .filter(CentroCustoModel.ug_id == data["ug_id"]) 
        .order_by(CentroCustoModel.id.desc())
        .first()
    )
    sequencial = 1
    if ultimo and ultimo.codigo:
        try:
            sequencial = int(ultimo.codigo.split('.')[-1]) + 1
        except Exception:
            sequencial = 1
    codigo = f"{ug.codigo}.{sequencial:03d}"

    c = CentroCustoModel(
        codigo=codigo,
        nome=data["nome"],
        empresa_id=data["empresa_id"],
        ug_id=data["ug_id"],
        responsavel_id=data.get("responsavel_id"),
        observacoes=data.get("observacoes"),
        criado_por=data.get("criado_por"),
        status=data.get("status", "Ativo"),
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

# -----------------------------
# Classes Contábeis - Schemas e CRUD
# -----------------------------
class ClasseContabil(BaseModel):
    id: int
    codigo: str
    descricao: str
    vida_util_anos: int
    taxa_depreciacao: float
    empresa_id: int
    conta_contabil_id: Optional[int] = None
    status: str
    criado_em: datetime

    class Config:
        from_attributes = True

class ClasseContabilCreate(BaseModel):
    codigo: str
    descricao: str
    vida_util_anos: int
    taxa_depreciacao: float
    empresa_id: int
    conta_contabil_id: Optional[int] = None
    status: str = "Ativo"

class ClasseContabilUpdate(BaseModel):
    codigo: Optional[str] = None
    descricao: Optional[str] = None
    vida_util_anos: Optional[int] = None
    taxa_depreciacao: Optional[float] = None
    empresa_id: Optional[int] = None
    conta_contabil_id: Optional[int] = None
    status: Optional[str] = None

@app.get("/classes_contabeis", response_model=List[ClasseContabil])
def list_classes_contabeis(
    empresa_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: UsuarioModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    q = db.query(ClasseContabilModel)
    allowed = get_allowed_company_ids(db, current_user)
    if allowed:
        q = q.filter(ClasseContabilModel.empresa_id.in_(allowed))
    if empresa_id:
        if allowed and empresa_id not in allowed:
            raise HTTPException(status_code=403, detail="Acesso negado à empresa informada")
        q = q.filter(ClasseContabilModel.empresa_id == empresa_id)
    if status:
        q = q.filter(ClasseContabilModel.status == status)
    return q.order_by(ClasseContabilModel.codigo).all()

@app.get("/classes_contabeis/{id}", response_model=ClasseContabil)
def get_classe_contabil(id: int, db: Session = Depends(get_db)):
    c = db.query(ClasseContabilModel).filter(ClasseContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Classe Contábil não encontrada")
    return c

@app.post("/classes_contabeis", response_model=ClasseContabil)
def create_classe_contabil(payload: ClasseContabilCreate, db: Session = Depends(get_db)):
    data = payload.dict()
    
    # Valida empresa
    if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
        raise HTTPException(status_code=400, detail="Empresa inválida")

    # Valida Conta Contábil se informada
    if data.get("conta_contabil_id"):
        cc = db.query(ContaContabilModel).filter(ContaContabilModel.id == data["conta_contabil_id"]).first()
        if not cc:
            raise HTTPException(status_code=400, detail="Conta Contábil não encontrada")
        if cc.empresa_id != data["empresa_id"]:
            raise HTTPException(status_code=400, detail="Conta Contábil pertence a outra empresa")
        
    # Check duplicate code in company
    if db.query(ClasseContabilModel).filter(ClasseContabilModel.codigo == data["codigo"], ClasseContabilModel.empresa_id == data["empresa_id"]).first():
        raise HTTPException(status_code=400, detail="Já existe uma Classe Contábil com este código nesta empresa")

    c = ClasseContabilModel(**data)
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/classes_contabeis/{id}", response_model=ClasseContabil)
def update_classe_contabil(id: int, payload: ClasseContabilUpdate, db: Session = Depends(get_db)):
    c = db.query(ClasseContabilModel).filter(ClasseContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Classe Contábil não encontrada")
    
    data = payload.dict(exclude_unset=True)
    
    if "codigo" in data:
        existing = db.query(ClasseContabilModel).filter(
            ClasseContabilModel.codigo == data["codigo"], 
            ClasseContabilModel.empresa_id == (data.get("empresa_id") or c.empresa_id),
            ClasseContabilModel.id != id
        ).first()
        if existing:
             raise HTTPException(status_code=400, detail="Já existe uma Classe Contábil com este código nesta empresa")

    for key, value in data.items():
        setattr(c, key, value)
        
    db.commit()
    db.refresh(c)
    return c

@app.delete("/classes_contabeis/{id}")
def delete_classe_contabil(id: int, db: Session = Depends(get_db)):
    c = db.query(ClasseContabilModel).filter(ClasseContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Classe Contábil não encontrada")
    db.delete(c)
    db.commit()
    return {"deleted": True}

# -----------------------------
# Contas Contábeis - Schemas e CRUD
# -----------------------------
class ContaContabil(BaseModel):
    id: int
    codigo: str
    descricao: str
    empresa_id: int
    status: str
    criado_em: datetime

    class Config:
        from_attributes = True

class ContaContabilCreate(BaseModel):
    codigo: str
    descricao: str
    empresa_id: int
    status: str = "Ativo"

class ContaContabilUpdate(BaseModel):
    codigo: Optional[str] = None
    descricao: Optional[str] = None
    empresa_id: Optional[int] = None
    status: Optional[str] = None

@app.get("/contas_contabeis", response_model=List[ContaContabil])
def list_contas_contabeis(
    empresa_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: UsuarioModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    q = db.query(ContaContabilModel)
    allowed = get_allowed_company_ids(db, current_user)
    if allowed:
        q = q.filter(ContaContabilModel.empresa_id.in_(allowed))
    if empresa_id:
         if allowed and empresa_id not in allowed:
            raise HTTPException(status_code=403, detail="Acesso negado à empresa informada")
         q = q.filter(ContaContabilModel.empresa_id == empresa_id)
    if status:
        q = q.filter(ContaContabilModel.status == status)
    return q.order_by(ContaContabilModel.codigo).all()

@app.get("/contas_contabeis/{id}", response_model=ContaContabil)
def get_conta_contabil(id: int, db: Session = Depends(get_db)):
    c = db.query(ContaContabilModel).filter(ContaContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conta Contábil não encontrada")
    return c

@app.post("/contas_contabeis", response_model=ContaContabil)
def create_conta_contabil(payload: ContaContabilCreate, db: Session = Depends(get_db)):
    data = payload.dict()
    
    # Valida empresa
    if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
        raise HTTPException(status_code=400, detail="Empresa inválida")

    # Check duplicate code in company
    if db.query(ContaContabilModel).filter(ContaContabilModel.codigo == data["codigo"], ContaContabilModel.empresa_id == data["empresa_id"]).first():
        raise HTTPException(status_code=400, detail="Já existe uma Conta Contábil com este código nesta empresa")

    c = ContaContabilModel(**data)
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/contas_contabeis/{id}", response_model=ContaContabil)
def update_conta_contabil(id: int, payload: ContaContabilUpdate, db: Session = Depends(get_db)):
    c = db.query(ContaContabilModel).filter(ContaContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conta Contábil não encontrada")
    
    data = payload.dict(exclude_unset=True)
    
    if "codigo" in data:
        existing = db.query(ContaContabilModel).filter(
            ContaContabilModel.codigo == data["codigo"], 
            ContaContabilModel.empresa_id == (data.get("empresa_id") or c.empresa_id),
            ContaContabilModel.id != id
        ).first()
        if existing:
             raise HTTPException(status_code=400, detail="Já existe uma Conta Contábil com este código nesta empresa")

    for key, value in data.items():
        setattr(c, key, value)
        
    db.commit()
    db.refresh(c)
    return c

@app.delete("/contas_contabeis/{id}")
def delete_conta_contabil(id: int, db: Session = Depends(get_db)):
    c = db.query(ContaContabilModel).filter(ContaContabilModel.id == id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conta Contábil não encontrada")
    db.delete(c)
    db.commit()
    return {"deleted": True}

# -----------------------------
# Usuários - Schemas e CRUD
# -----------------------------
class Usuario(BaseModel):
    id: int
    codigo: str
    nome_completo: str
    email: EmailStr
    cpf: str
    nome_usuario: str
    data_nascimento: Optional[date] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    status: str
    class Config:
        from_attributes = True

class UsuarioCreate(BaseModel):
    nome_completo: str
    email: EmailStr
    senha: str
    confirmacao_senha: str
    cpf: str
    nome_usuario: str
    data_nascimento: Optional[date] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    status: str = "Ativo"

class UsuarioUpdate(BaseModel):
    nome_completo: Optional[str] = None
    email: Optional[EmailStr] = None
    senha: Optional[str] = None
    confirmacao_senha: Optional[str] = None
    cpf: Optional[str] = None
    nome_usuario: Optional[str] = None
    data_nascimento: Optional[date] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    status: Optional[str] = None


def _generate_user_codigo(db: Session) -> str:
    last = db.query(UsuarioModel.id).order_by(UsuarioModel.id.desc()).first()
    next_num = (last[0] if last else 0) + 1
    return f"{next_num:06d}"


@app.get("/usuarios", response_model=List[Usuario])
def list_usuarios(request: Request, db: Session = Depends(get_db)):
    empresa_hdr = request.headers.get("X-Company-Id")
    q = db.query(UsuarioModel)
    if empresa_hdr:
        try:
            empresa_id = int(empresa_hdr)
            q = q.filter(sa.or_(UsuarioModel.empresa_id == empresa_id, UsuarioModel.empresa_id == None))
        except ValueError:
            pass
    return q.order_by(UsuarioModel.id).all()


@app.get("/usuarios/{user_id}", response_model=Usuario)
def get_usuario(user_id: int, db: Session = Depends(get_db)):
    u = db.query(UsuarioModel).filter(UsuarioModel.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return u


@app.post("/usuarios", response_model=Usuario)
def create_usuario(payload: UsuarioCreate, db: Session = Depends(get_db)):
    if payload.senha != payload.confirmacao_senha:
        raise HTTPException(status_code=400, detail="Confirmação de senha não confere")
    # checa duplicidades
    if db.query(UsuarioModel).filter(UsuarioModel.email == payload.email).first():
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    if db.query(UsuarioModel).filter(UsuarioModel.cpf == payload.cpf).first():
        raise HTTPException(status_code=400, detail="CPF já cadastrado")
    if db.query(UsuarioModel).filter(UsuarioModel.nome_usuario == payload.nome_usuario).first():
        raise HTTPException(status_code=400, detail="Nome de usuário já cadastrado")
    # valida FKs se enviados
    if payload.empresa_id is not None:
        if not db.query(CompanyModel).filter(CompanyModel.id == payload.empresa_id).first():
            raise HTTPException(status_code=400, detail="Empresa inválida")
    if payload.ug_id is not None:
        if not db.query(UGModel).filter(UGModel.id == payload.ug_id).first():
            raise HTTPException(status_code=400, detail="UG inválida")

    senha_hash = bcrypt.hashpw(payload.senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    codigo = _generate_user_codigo(db)
    u = UsuarioModel(
        codigo=codigo,
        nome_completo=payload.nome_completo,
        email=payload.email,
        senha_hash=senha_hash,
        cpf=payload.cpf,
        nome_usuario=payload.nome_usuario,
        data_nascimento=payload.data_nascimento,
        empresa_id=payload.empresa_id,
        ug_id=payload.ug_id,
        centro_custo_id=payload.centro_custo_id,
        status=payload.status,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return u


@app.put("/usuarios/{user_id}", response_model=Usuario)
def update_usuario(user_id: int, payload: UsuarioUpdate, db: Session = Depends(get_db)):
    u = db.query(UsuarioModel).filter(UsuarioModel.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    data = payload.dict(exclude_unset=True)

    # valida duplicidades se alterando
    if "email" in data:
        exists = db.query(UsuarioModel).filter(UsuarioModel.email == data["email"], UsuarioModel.id != user_id).first()
        if exists:
            raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    if "cpf" in data:
        exists = db.query(UsuarioModel).filter(UsuarioModel.cpf == data["cpf"], UsuarioModel.id != user_id).first()
        if exists:
            raise HTTPException(status_code=400, detail="CPF já cadastrado")
    if "nome_usuario" in data:
        exists = db.query(UsuarioModel).filter(UsuarioModel.nome_usuario == data["nome_usuario"], UsuarioModel.id != user_id).first()
        if exists:
            raise HTTPException(status_code=400, detail="Nome de usuário já cadastrado")

    # valida FKs
    if "empresa_id" in data and data["empresa_id"] is not None:
        if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
            raise HTTPException(status_code=400, detail="Empresa inválida")
    if "ug_id" in data and data["ug_id"] is not None:
        if not db.query(UGModel).filter(UGModel.id == data["ug_id"]).first():
            raise HTTPException(status_code=400, detail="UG inválida")

    # tratar senha
    if data.get("senha") or data.get("confirmacao_senha"):
        if data.get("senha") != data.get("confirmacao_senha"):
            raise HTTPException(status_code=400, detail="Confirmação de senha não confere")
        u.senha_hash = bcrypt.hashpw(data["senha"].encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        # remove campos transitórios
        data.pop("senha", None)
        data.pop("confirmacao_senha", None)

    for k, v in data.items():
        setattr(u, k, v)

    db.commit()
    db.refresh(u)
    return u


@app.delete("/usuarios/{user_id}")
def delete_usuario(user_id: int, db: Session = Depends(get_db)):
    u = db.query(UsuarioModel).filter(UsuarioModel.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    db.delete(u)
    db.commit()
    return {"deleted": True}

class RevisaoPeriodo(BaseModel):
    id: int
    codigo: str
    descricao: str
    data_abertura: date
    data_fechamento_prevista: date
    data_fechamento: Optional[date] = None
    data_inicio_nova_vida_util: Optional[date] = None
    empresa_id: int
    ug_id: Optional[int] = None
    responsavel_id: int
    status: str
    observacoes: Optional[str] = None
    criado_em: datetime
    class Config:
        from_attributes = True

class RevisaoPeriodoCreate(BaseModel):
    descricao: str
    data_abertura: date
    data_fechamento_prevista: date
    data_inicio_nova_vida_util: Optional[date] = None
    empresa_id: int
    ug_id: Optional[int] = None
    responsavel_id: int
    status: str = "Aberto"
    observacoes: Optional[str] = None

class RevisaoPeriodoUpdate(BaseModel):
    descricao: Optional[str] = None
    data_abertura: Optional[date] = None
    data_fechamento_prevista: Optional[date] = None
    data_fechamento: Optional[date] = None
    data_inicio_nova_vida_util: Optional[date] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    responsavel_id: Optional[int] = None
    status: Optional[str] = None
    observacoes: Optional[str] = None

class RevisaoItemOut(BaseModel):
    id: int
    periodo_id: int
    numero_imobilizado: str
    sub_numero: str
    descricao: str
    data_inicio_depreciacao: date
    data_fim_depreciacao: Optional[date] = None
    data_fim_revisada: Optional[date] = None
    valor_aquisicao: float
    depreciacao_acumulada: float
    valor_contabil: float
    centro_custo: str
    classe: str
    descricao_classe: Optional[str] = None
    conta_contabil: str
    descricao_conta_contabil: str
    vida_util_anos: int
    vida_util_periodos: int
    vida_util_revisada: Optional[int] = None
    condicao_fisica: Optional[str] = None
    justificativa: Optional[str] = None
    alterado: Optional[bool] = None
    auxiliar2: Optional[str] = None
    auxiliar3: Optional[str] = None
    status: str
    criado_em: datetime
    class Config:
        from_attributes = True

ALLOWED_STATUS = {"Aberto", "Em Andamento", "Fechado"}

def _generate_revisao_codigo(db: Session, ano: int) -> str:
    prefix = f"RV{ano}-"
    # conta existentes do ano
    count = (
        db.query(RevisaoPeriodoModel)
        .filter(RevisaoPeriodoModel.codigo.like(f"{prefix}%"))
        .count()
    )
    return f"{prefix}{count + 1:02d}"

@app.get("/revisoes/periodos", response_model=List[RevisaoPeriodo])
def list_revisoes(db: Session = Depends(get_db), current_user: UsuarioModel = Depends(get_current_user)):
    allowed = get_allowed_company_ids(db, current_user)
    if not allowed:
        return []
        
    q = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.empresa_id.in_(allowed))
    
    return q.order_by(RevisaoPeriodoModel.id.desc()).all()

@app.post("/revisoes/periodos", response_model=RevisaoPeriodo)
def create_revisao(payload: RevisaoPeriodoCreate, db: Session = Depends(get_db)):
    if payload.status not in ALLOWED_STATUS:
        raise HTTPException(status_code=400, detail="Status inválido")
    # valida responsável
    if not db.query(UsuarioModel).filter(UsuarioModel.id == payload.responsavel_id).first():
        raise HTTPException(status_code=400, detail="Responsável inválido")
    # valida empresa e obtém dados para descrição
    company = db.query(CompanyModel).filter(CompanyModel.id == payload.empresa_id).first()
    if not company:
        raise HTTPException(status_code=400, detail="Empresa inválida")
            
    ano = payload.data_abertura.year
    
    # Verificar se existe algum período em aberto para esta empresa (não fechado)
    previous_open = db.query(RevisaoPeriodoModel).filter(
        RevisaoPeriodoModel.empresa_id == payload.empresa_id,
        RevisaoPeriodoModel.status != 'Fechado'
    ).first()
    
    if previous_open:
        raise HTTPException(status_code=400, detail=f"Existe um período anterior ({previous_open.codigo}) que não está fechado. Encerre-o antes de abrir um novo.")

    # Verificar se já existe período para esta empresa neste ano
    existing = db.query(RevisaoPeriodoModel).filter(
        RevisaoPeriodoModel.empresa_id == payload.empresa_id,
        sa.extract('year', RevisaoPeriodoModel.data_abertura) == ano
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail=f"O sistema só permite um período por ano para cada empresa. Já existe o período {existing.codigo} para o ano {ano}.")

    codigo = _generate_revisao_codigo(db, ano)
    
    # Gerar descrição padrão se não fornecida ou se for vazia
    descricao_final = payload.descricao
    if not descricao_final:
        cidade = f" - {company.city}" if company.city else ""
        descricao_final = f"{company.name}{cidade}"
    
    r = RevisaoPeriodoModel(
        codigo=codigo,
        descricao=descricao_final,
        data_abertura=payload.data_abertura,
        data_fechamento_prevista=payload.data_fechamento_prevista,
        data_inicio_nova_vida_util=payload.data_inicio_nova_vida_util,
        empresa_id=payload.empresa_id,
        ug_id=payload.ug_id,
        responsavel_id=payload.responsavel_id,
        status=payload.status,
        observacoes=payload.observacoes,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return r

@app.put("/revisoes/periodos/{rev_id}", response_model=RevisaoPeriodo)
def update_revisao(rev_id: int, payload: RevisaoPeriodoUpdate, db: Session = Depends(get_db)):
    r = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Período não encontrado")
    data = payload.dict(exclude_unset=True)
    if "status" in data and data["status"] not in ALLOWED_STATUS:
        raise HTTPException(status_code=400, detail="Status inválido")
    if "responsavel_id" in data and data["responsavel_id"] is not None:
        if not db.query(UsuarioModel).filter(UsuarioModel.id == data["responsavel_id"]).first():
            raise HTTPException(status_code=400, detail="Responsável inválido")
    if "empresa_id" in data and data["empresa_id"] is not None:
        if not db.query(CompanyModel).filter(CompanyModel.id == data["empresa_id"]).first():
            raise HTTPException(status_code=400, detail="Empresa inválida")
    for k, v in data.items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    return r

@app.delete("/revisoes/periodos/{rev_id}")
def delete_revisao(rev_id: int, db: Session = Depends(get_db)):
    r = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Período não encontrado")
    db.delete(r)
    db.commit()
    return {"deleted": True}

# ========================================
# Controle de Acessos - Permissões por Grupo
# ========================================

class Transacao(BaseModel):
    id: int
    nome_tela: str
    rota: str
    descricao: Optional[str] = None
    class Config:
        from_attributes = True

class TransacaoCreate(BaseModel):
    nome_tela: str
    rota: str
    descricao: Optional[str] = None

class TransacaoUpdate(BaseModel):
    nome_tela: Optional[str] = None
    rota: Optional[str] = None
    descricao: Optional[str] = None

@app.get("/permissoes/transacoes", response_model=List[Transacao])
def list_transacoes(q: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(TransacaoModel)
    if q:
        query = query.filter(sa.func.lower(TransacaoModel.nome_tela).like(f"%{q.lower()}%"))
    return query.order_by(TransacaoModel.nome_tela).all()

@app.post("/permissoes/transacoes", response_model=Transacao)
def create_transacao(payload: TransacaoCreate, db: Session = Depends(get_db)):
    # rota única
    if db.query(TransacaoModel).filter(TransacaoModel.rota == payload.rota).first():
        raise HTTPException(status_code=400, detail="Rota já cadastrada")
    t = TransacaoModel(**payload.dict())
    db.add(t)
    db.commit()
    db.refresh(t)
    audit(db, usuario_id=None, acao="create", entidade="transacao", entidade_id=t.id, detalhes=f"rota={t.rota}")
    return t

@app.put("/permissoes/transacoes/{transacao_id}", response_model=Transacao)
def update_transacao(transacao_id: int, payload: TransacaoUpdate, db: Session = Depends(get_db)):
    t = db.query(TransacaoModel).filter(TransacaoModel.id == transacao_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    data = payload.dict(exclude_unset=True)
    if "rota" in data:
        exists = db.query(TransacaoModel).filter(TransacaoModel.rota == data["rota"], TransacaoModel.id != transacao_id).first()
        if exists:
            raise HTTPException(status_code=400, detail="Rota já cadastrada")
    for k, v in data.items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    audit(db, usuario_id=None, acao="update", entidade="transacao", entidade_id=t.id)
    return t

@app.delete("/permissoes/transacoes/{transacao_id}")
def delete_transacao(transacao_id: int, db: Session = Depends(get_db)):
    t = db.query(TransacaoModel).filter(TransacaoModel.id == transacao_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    db.delete(t)
    db.commit()
    audit(db, usuario_id=None, acao="delete", entidade="transacao", entidade_id=transacao_id)
    return {"deleted": True}


class GrupoPermissao(BaseModel):
    id: int
    nome: str
    descricao: Optional[str] = None
    criado_em: datetime
    atualizado_em: datetime
    class Config:
        from_attributes = True

class GrupoPermissaoCreate(BaseModel):
    nome: str
    descricao: Optional[str] = None

class GrupoPermissaoUpdate(BaseModel):
    nome: Optional[str] = None
    descricao: Optional[str] = None

class GrupoResumo(BaseModel):
    id: int
    nome: str
    descricao: Optional[str] = None
    usuarios: int
    empresas: int
    transacoes: int

@app.get("/permissoes/grupos", response_model=List[GrupoResumo])
def list_grupos(q: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(GrupoPermissaoModel)
    if q:
        query = query.filter(sa.func.lower(GrupoPermissaoModel.nome).like(f"%{q.lower()}%"))
    grupos = query.order_by(GrupoPermissaoModel.nome).all()
    result: List[GrupoResumo] = []
    for g in grupos:
        result.append(GrupoResumo(
            id=g.id,
            nome=g.nome,
            descricao=g.descricao,
            usuarios=db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.grupo_id == g.id).count(),
            empresas=db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id == g.id).count(),
            transacoes=db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id == g.id).count(),
        ))
    return result

@app.get("/permissoes/grupos/{grupo_id}", response_model=GrupoPermissao)
def get_grupo(grupo_id: int, db: Session = Depends(get_db)):
    g = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    return g

@app.post("/permissoes/grupos", response_model=GrupoPermissao)
def create_grupo(payload: GrupoPermissaoCreate, db: Session = Depends(get_db)):
    if db.query(GrupoPermissaoModel).filter(sa.func.lower(GrupoPermissaoModel.nome) == payload.nome.lower()).first():
        raise HTTPException(status_code=400, detail="Nome do grupo já existe")
    g = GrupoPermissaoModel(**payload.dict())
    db.add(g)
    db.commit()
    db.refresh(g)
    audit(db, usuario_id=None, acao="create", entidade="grupo", entidade_id=g.id, detalhes=f"nome={g.nome}")
    return g

@app.put("/permissoes/grupos/{grupo_id}", response_model=GrupoPermissao)
def update_grupo(grupo_id: int, payload: GrupoPermissaoUpdate, db: Session = Depends(get_db)):
    g = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    data = payload.dict(exclude_unset=True)
    if "nome" in data:
        exists = db.query(GrupoPermissaoModel).filter(sa.func.lower(GrupoPermissaoModel.nome) == data["nome"].lower(), GrupoPermissaoModel.id != grupo_id).first()
        if exists:
            raise HTTPException(status_code=400, detail="Nome do grupo já existe")
    for k, v in data.items():
        setattr(g, k, v)
    db.commit()
    db.refresh(g)
    audit(db, usuario_id=None, acao="update", entidade="grupo", entidade_id=g.id)
    return g

@app.delete("/permissoes/grupos/{grupo_id}")
def delete_grupo(grupo_id: int, db: Session = Depends(get_db)):
    g = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    # remove associações explícitas (garantido por cascade, mas fazemos por segurança)
    db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id == grupo_id).delete()
    db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id == grupo_id).delete()
    db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.grupo_id == grupo_id).delete()
    db.delete(g)
    db.commit()
    audit(db, usuario_id=None, acao="delete", entidade="grupo", entidade_id=grupo_id)
    return {"deleted": True}

# --- Associações ---
class IdRef(BaseModel):
    id: int

@app.get("/permissoes/grupos/{grupo_id}/empresas", response_model=List[Company])
def list_grupo_empresas(grupo_id: int, db: Session = Depends(get_db)):
    links = db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id == grupo_id).all()
    ids = [l.empresa_id for l in links]
    return db.query(CompanyModel).filter(CompanyModel.id.in_(ids)).order_by(CompanyModel.name).all()

@app.post("/permissoes/grupos/{grupo_id}/empresas")
def add_grupo_empresa(grupo_id: int, payload: IdRef, db: Session = Depends(get_db)):
    if not db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first():
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if not db.query(CompanyModel).filter(CompanyModel.id == payload.id).first():
        raise HTTPException(status_code=400, detail="Empresa inválida")
    link = GrupoEmpresaModel(grupo_id=grupo_id, empresa_id=payload.id)
    try:
        db.add(link)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Associação já existe")
    audit(db, usuario_id=None, acao="link", entidade="grupo_empresa", entidade_id=link.id)
    return {"linked": True}

@app.delete("/permissoes/grupos/{grupo_id}/empresas/{empresa_id}")
def remove_grupo_empresa(grupo_id: int, empresa_id: int, db: Session = Depends(get_db)):
    deleted = db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id == grupo_id, GrupoEmpresaModel.empresa_id == empresa_id).delete()
    db.commit()
    return {"deleted": bool(deleted)}

@app.get("/permissoes/grupos/{grupo_id}/transacoes", response_model=List[Transacao])
def list_grupo_transacoes(grupo_id: int, db: Session = Depends(get_db)):
    links = db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id == grupo_id).all()
    ids = [l.transacao_id for l in links]
    return db.query(TransacaoModel).filter(TransacaoModel.id.in_(ids)).order_by(TransacaoModel.nome_tela).all()

@app.post("/permissoes/grupos/{grupo_id}/transacoes")
def add_grupo_transacao(grupo_id: int, payload: IdRef, db: Session = Depends(get_db)):
    if not db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first():
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if not db.query(TransacaoModel).filter(TransacaoModel.id == payload.id).first():
        raise HTTPException(status_code=400, detail="Transação inválida")
    link = GrupoTransacaoModel(grupo_id=grupo_id, transacao_id=payload.id)
    try:
        db.add(link)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Associação já existe")
    audit(db, usuario_id=None, acao="link", entidade="grupo_transacao", entidade_id=link.id)
    return {"linked": True}

@app.delete("/permissoes/grupos/{grupo_id}/transacoes/{transacao_id}")
def remove_grupo_transacao(grupo_id: int, transacao_id: int, db: Session = Depends(get_db)):
    deleted = db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id == grupo_id, GrupoTransacaoModel.transacao_id == transacao_id).delete()
    db.commit()
    return {"deleted": bool(deleted)}

class Usuario(BaseModel):
    id: int
    codigo: str
    nome_completo: str
    email: EmailStr
    cpf: str
    nome_usuario: str
    data_nascimento: Optional[date] = None
    empresa_id: Optional[int] = None
    ug_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    status: str
    class Config:
        from_attributes = True

@app.get("/permissoes/grupos/{grupo_id}/usuarios", response_model=List[Usuario])
def list_grupo_usuarios(grupo_id: int, db: Session = Depends(get_db)):
    links = db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.grupo_id == grupo_id).all()
    ids = [l.usuario_id for l in links]
    return db.query(UsuarioModel).filter(UsuarioModel.id.in_(ids)).order_by(UsuarioModel.nome_completo).all()

@app.post("/permissoes/grupos/{grupo_id}/usuarios")
def add_grupo_usuario(grupo_id: int, payload: IdRef, db: Session = Depends(get_db)):
    if not db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first():
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if not db.query(UsuarioModel).filter(UsuarioModel.id == payload.id).first():
        raise HTTPException(status_code=400, detail="Usuário inválido")
    link = GrupoUsuarioModel(grupo_id=grupo_id, usuario_id=payload.id)
    try:
        db.add(link)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Associação já existe")
    audit(db, usuario_id=None, acao="link", entidade="grupo_usuario", entidade_id=link.id)
    return {"linked": True}

@app.delete("/permissoes/grupos/{grupo_id}/usuarios/{usuario_id}")
def remove_grupo_usuario(grupo_id: int, usuario_id: int, db: Session = Depends(get_db)):
    deleted = db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.grupo_id == grupo_id, GrupoUsuarioModel.usuario_id == usuario_id).delete()
    db.commit()
    return {"deleted": bool(deleted)}


# --- Clonagem de Grupo ---
class ClonePayload(BaseModel):
    novo_nome: str
    descricao: Optional[str] = None

@app.post("/permissoes/grupos/{grupo_id}/clonar", response_model=GrupoPermissao)
def clonar_grupo(grupo_id: int, payload: ClonePayload, db: Session = Depends(get_db)):
    orig = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id == grupo_id).first()
    if not orig:
        raise HTTPException(status_code=404, detail="Grupo de origem não encontrado")
    if db.query(GrupoPermissaoModel).filter(sa.func.lower(GrupoPermissaoModel.nome) == payload.novo_nome.lower()).first():
        raise HTTPException(status_code=400, detail="Nome do grupo já existe")
    novo = GrupoPermissaoModel(nome=payload.novo_nome, descricao=payload.descricao or orig.descricao)
    db.add(novo)
    db.flush()
    # copiar vínculos
    for ge in db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id == grupo_id).all():
        db.add(GrupoEmpresaModel(grupo_id=novo.id, empresa_id=ge.empresa_id))
    for gt in db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id == grupo_id).all():
        db.add(GrupoTransacaoModel(grupo_id=novo.id, transacao_id=gt.transacao_id))
    for gu in db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.grupo_id == grupo_id).all():
        db.add(GrupoUsuarioModel(grupo_id=novo.id, usuario_id=gu.usuario_id))
    db.commit()
    db.refresh(novo)
    audit(db, usuario_id=None, acao="clone", entidade="grupo", entidade_id=novo.id, detalhes=f"from={grupo_id}")
    return novo

# --- Consulta permissões de usuário ---
class UsuarioPermissoes(BaseModel):
    usuario_id: int
    grupos: List[str]
    empresas_ids: List[int]
    rotas: List[str]

@app.get("/permissoes/usuario/{usuario_id}", response_model=UsuarioPermissoes)
def permissoes_usuario(usuario_id: int, db: Session = Depends(get_db)):
    if not db.query(UsuarioModel).filter(UsuarioModel.id == usuario_id).first():
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    links = db.query(GrupoUsuarioModel).filter(GrupoUsuarioModel.usuario_id == usuario_id).all()
    grupo_ids = [l.grupo_id for l in links]
    grupos = db.query(GrupoPermissaoModel).filter(GrupoPermissaoModel.id.in_(grupo_ids)).all()
    empresas = db.query(GrupoEmpresaModel).filter(GrupoEmpresaModel.grupo_id.in_(grupo_ids)).all()
    transacoes = db.query(GrupoTransacaoModel).filter(GrupoTransacaoModel.grupo_id.in_(grupo_ids)).all()
    rotas = []
    if transacoes:
        t_ids = [t.transacao_id for t in transacoes]
        rotas = [t.rota for t in db.query(TransacaoModel).filter(TransacaoModel.id.in_(t_ids)).all()]
    return UsuarioPermissoes(
        usuario_id=usuario_id,
        grupos=[g.nome for g in grupos],
        empresas_ids=[e.empresa_id for e in empresas],
        rotas=rotas,
    )

@app.post("/revisoes/fechar/{rev_id}", response_model=RevisaoPeriodo)
def fechar_revisao(rev_id: int, db: Session = Depends(get_db)):
    r = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Período não encontrado")
    
    # Validar se o cronograma (se existir) está concluído
    c = db.query(CronogramaModel).filter(CronogramaModel.periodo_id == rev_id).first()
    if c:
        pending = db.execute(sa.text(
            "SELECT COUNT(*) FROM cronogramas_tarefas WHERE cronograma_id = :cid AND status != 'Concluída'"
        ), {"cid": c.id}).scalar()
        if pending > 0:
            raise HTTPException(status_code=400, detail="O cronograma deste período possui tarefas pendentes. Conclua o cronograma antes de fechar o período.")

    # Validar se todos os ativos foram delegados
    total_items = db.query(RevisaoItemModel).filter(RevisaoItemModel.periodo_id == rev_id).count()
    delegated_items = db.query(RevisaoDelegacaoModel.ativo_id).filter(RevisaoDelegacaoModel.periodo_id == rev_id).distinct().count()

    if total_items > delegated_items:
        diff = total_items - delegated_items
        raise HTTPException(status_code=400, detail=f"Existem {diff} ativos não delegados. De acordo com a norma, todos os ativos devem ser revisados (delegados) antes do encerramento.")

    r.status = "Fechado"
    r.data_fechamento = date.today()
    db.commit()
    db.refresh(r)
    return r

@app.post("/revisoes/upload_base/{rev_id}")
def upload_base(rev_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    r = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Período não encontrado")
    if r.status == "Fechado":
        raise HTTPException(status_code=400, detail="Período está fechado")

    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Formato inválido. Envie .csv ou .xlsx")

    required = {
        "numero_imobilizado",
        "sub_numero",
        "data_inicio_depreciacao",
        "descricao",
        "valor_aquisicao",
        "depreciacao_acumulada",
        "valor_contabil",
        "centro_custo",
        "classe",
        "conta_contabil",
        "descricao_conta_contabil",
        "vida_util_anos",
        "vida_util_periodos",
    }
    optional = {"data_fim_depreciacao", "auxiliar2", "auxiliar3", "descricao_classe"}

    def norm_key(k: str) -> str:
        import re, unicodedata
        def clean(s: str) -> str:
            s = (s or "").strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')  # remove acentos
            s = s.replace('º', 'o').replace('ª', 'a')
            s = re.sub(r'[^a-z0-9]+', '_', s)  # troca não alfanum por _
            s = re.sub(r'_+', '_', s).strip('_')
            return s
        nk = clean(k)
        synonyms = {
            'no_imobilizado': 'numero_imobilizado',
            'n_imobilizado': 'numero_imobilizado',
            'numero_do_imobilizado': 'numero_imobilizado',
            'numero_imobilizado': 'numero_imobilizado',
            'sub_no': 'sub_numero',
            'sub_numero': 'sub_numero',
            'data_inicio_da_depreciacao': 'data_inicio_depreciacao',
            'data_inicio_de_depreciacao': 'data_inicio_depreciacao',
            'data_inicio_depreciacao': 'data_inicio_depreciacao',
            'descricao': 'descricao',
            'valor_aquisicao': 'valor_aquisicao',
            'depreciacao_acum': 'depreciacao_acumulada',
            'depreciacao_acumulada': 'depreciacao_acumulada',
            'valor_contabil': 'valor_contabil',
            'centro_custos': 'centro_custo',
            'centro_de_custos': 'centro_custo',
            'centro_custo': 'centro_custo',
            'classe': 'classe',
            'desc_classe': 'descricao_classe',
            'descricao_da_classe': 'descricao_classe',
            'descricao_classe': 'descricao_classe',
            'conta_contabil': 'conta_contabil',
            'desc_conta_contabil': 'descricao_conta_contabil',
            'descricao_conta_contabil': 'descricao_conta_contabil',
            'vida_util_anos': 'vida_util_anos',
            'vida_util_periodos': 'vida_util_periodos',
            'data_fim_depreciacao': 'data_fim_depreciacao',
            # auxiliares opcionais
            'auxiliar_1': 'auxiliar2',
            'aux_1': 'auxiliar2',
            'auxiliar1': 'auxiliar2',
            'auxiliar_2': 'auxiliar3',
            'aux_2': 'auxiliar3',
            'auxiliar2': 'auxiliar2',
            'auxiliar3': 'auxiliar3',
        }
        return synonyms.get(nk, nk)

    from decimal import Decimal
    import csv, io
    from datetime import datetime as dt
    import calendar

    def parse_date_any(x) -> date:
        if isinstance(x, date):
            return x
        s = str(x or "").strip()
        if not s:
            return None
        # Excel serial date (1900 date system): number of days since 1899-12-30
        try:
            n = float(s)
            if n > 20000 and n < 80000:
                from datetime import timedelta
                return date(1899, 12, 30) + timedelta(days=int(round(n)))
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return dt.strptime(s, fmt).date()
            except Exception:
                pass
        raise ValueError(f"Data inválida: {s}")

    def add_months(d: date, months: int) -> date:
        y = d.year + (d.month - 1 + months) // 12
        m = (d.month - 1 + months) % 12 + 1
        last_day = calendar.monthrange(y, m)[1]
        day = min(d.day, last_day)
        return date(y, m, day)

    imported = 0
    rejected = 0
    errors_log = []
    warnings_log = []

    try:
        content = file.file.read()
        rows = []
        headers = []
        if filename.endswith(".csv"):
            text = content.decode("utf-8", errors="replace")
            try:
                dialect = csv.Sniffer().sniff(text.splitlines()[0])
            except Exception:
                dialect = csv.excel
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            headers = [norm_key(h) for h in reader.fieldnames or []]
            for row in reader:
                rows.append({norm_key(k): v for k, v in row.items()})
        else:
            try:
                from openpyxl import load_workbook
            except Exception:
                raise HTTPException(status_code=500, detail="Dependência 'openpyxl' não instalada para XLSX")
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [norm_key(str(h)) for h in header_row]
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {}
                for idx, h in enumerate(headers):
                    row[h] = r[idx] if idx < len(r) else None
                rows.append(row)

        missing = [h for h in sorted(required) if h not in headers]
        if missing:
            raise HTTPException(status_code=400, detail=f"Cabeçalhos ausentes: {', '.join(missing)}")

        # Limpa itens anteriores (opcional). Comentado por segurança.
        # db.query(RevisaoItemModel).filter(RevisaoItemModel.periodo_id == rev_id).delete()

        for idx, row in enumerate(rows, start=2):  # start=2 considera header na linha 1
            try:
                numero_imobilizado = str(row.get("numero_imobilizado", "")).strip()
                sub_numero = str(row.get("sub_numero", "")).strip()
                descricao = str(row.get("descricao", "")).strip()
                if not (numero_imobilizado and sub_numero and descricao):
                    raise ValueError("Campos chave vazios")

                data_inicio = row.get("data_inicio_depreciacao")
                data_inicio = parse_date_any(data_inicio)

                def parse_int(x):
                    s = str(x).strip()
                    if s == "":
                        return 0
                    try:
                        return int(s)
                    except Exception:
                        try:
                            return int(float(s.replace(',', '.')))
                        except Exception:
                            raise ValueError(f"Valor inteiro inválido: {x}")

                vida_util_anos = parse_int(row.get("vida_util_anos"))
                vida_util_periodos = parse_int(row.get("vida_util_periodos"))
                if vida_util_periodos == 0 and vida_util_anos > 0:
                    vida_util_periodos = vida_util_anos * 12
                if vida_util_anos == 0 and vida_util_periodos > 0:
                    vida_util_anos = max(0, round(vida_util_periodos / 12))
                if str(row.get("vida_util_anos", "")).strip() == "":
                    warnings_log.append(f"Linha {idx}: vida_util_anos vazio")
                if str(row.get("vida_util_anos", "")).strip() == "" and str(row.get("vida_util_periodos", "")).strip() == "":
                    raise ValueError("Vida útil em branco: informe anos ou períodos")

                def parse_decimal(x):
                    x = (str(x).replace(".", "").replace(",", ".") if x is not None else "0")
                    return Decimal(x or "0")

                valor_aquisicao = parse_decimal(row.get("valor_aquisicao"))
                depreciacao_acumulada = parse_decimal(row.get("depreciacao_acumulada"))
                valor_contabil = parse_decimal(row.get("valor_contabil"))

                centro_custo = str(row.get("centro_custo", "")).strip()
                classe = str(row.get("classe", "")).strip()
                conta_contabil = str(row.get("conta_contabil", "")).strip()
                descricao_cc = str(row.get("descricao_conta_contabil", "")).strip()
                descricao_classe = str(row.get("descricao_classe", "")).strip()

                data_fim = row.get("data_fim_depreciacao")
                data_fim = parse_date_any(data_fim)
                if not data_fim and data_inicio and vida_util_periodos:
                    data_fim = add_months(data_inicio, vida_util_periodos)

                item = RevisaoItemModel(
                    periodo_id=rev_id,
                    numero_imobilizado=numero_imobilizado,
                    sub_numero=sub_numero,
                    descricao=descricao,
                    data_inicio_depreciacao=data_inicio,
                    data_fim_depreciacao=data_fim,
                    valor_aquisicao=valor_aquisicao,
                    depreciacao_acumulada=depreciacao_acumulada,
                    valor_contabil=valor_contabil,
                    centro_custo=centro_custo,
                    classe=classe,
                    descricao_classe=(descricao_classe or None),
                    conta_contabil=conta_contabil,
                    descricao_conta_contabil=descricao_cc,
                    vida_util_anos=vida_util_anos,
                    vida_util_periodos=vida_util_periodos,
                    auxiliar2=str(row.get("auxiliar2", "")) or None,
                    auxiliar3=str(row.get("auxiliar3", "")) or None,
                    status="Pendente",
                )
                db.add(item)
                imported += 1
            except Exception as ex:
                rejected += 1
                errors_log.append(f"Linha {idx}: {ex}")

        db.commit()
        return {
            "periodo_id": rev_id,
            "importados": imported,
            "rejeitados": rejected,
            "erros": errors_log,
            "alertas": warnings_log,
        }
    finally:
        try:
            file.file.close()
        except Exception:
            pass
@app.get("/revisoes/itens/{rev_id}", response_model=List[RevisaoItemOut])
def list_revisao_itens(rev_id: int, db: Session = Depends(get_db)):
    periodo = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail="Período de revisão não encontrado")

    items = (
        db.query(RevisaoItemModel)
        .filter(RevisaoItemModel.periodo_id == rev_id)
        .order_by(RevisaoItemModel.id.desc())
        .all()
    )

    # conversão Numeric -> float
    result = []
    for i in items:
        result.append(
            RevisaoItemOut(
                id=i.id,
                periodo_id=i.periodo_id,
                numero_imobilizado=i.numero_imobilizado,
                sub_numero=i.sub_numero,
                descricao=i.descricao,
                data_inicio_depreciacao=i.data_inicio_depreciacao,
                data_fim_depreciacao=i.data_fim_depreciacao,
                data_fim_revisada=i.data_fim_revisada,
                valor_aquisicao=float(i.valor_aquisicao) if i.valor_aquisicao is not None else 0.0,
                depreciacao_acumulada=float(i.depreciacao_acumulada) if i.depreciacao_acumulada is not None else 0.0,
                valor_contabil=float(i.valor_contabil) if i.valor_contabil is not None else 0.0,
                centro_custo=i.centro_custo,
                classe=i.classe,
                descricao_classe=getattr(i, 'descricao_classe', None),
                conta_contabil=i.conta_contabil,
                descricao_conta_contabil=i.descricao_conta_contabil,
                vida_util_anos=i.vida_util_anos,
                vida_util_periodos=i.vida_util_periodos,
                vida_util_revisada=i.vida_util_revisada,
                condicao_fisica=i.condicao_fisica,
                justificativa=i.justificativa,
                alterado=i.alterado,
                auxiliar2=i.auxiliar2,
                auxiliar3=i.auxiliar3,
                status=i.status,
                criado_em=i.criado_em,
            )
        )
    return result

# Atualização de item da revisão (vida útil revisada, condição física, justificativa)
class RevisaoItemUpdate(BaseModel):
    vida_util_revisada: Optional[int] = None
    condicao_fisica: Optional[str] = None  # Bom | Regular | Ruim
    justificativa: Optional[str] = None
    revisor_id: Optional[int] = None
    incremento: Optional[str] = None  # Acréscimo | Decréscimo | Manter
    motivo: Optional[str] = None
    nova_data_fim: Optional[date] = None

def _add_months(d: date, months: int) -> date:
    import calendar
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    last_day = calendar.monthrange(y, m)[1]
    day = min(d.day, last_day)
    return date(y, m, day)

# diferença em meses entre duas datas, ajustando pelo dia do mês
def _months_diff(start: date, end: date) -> int:
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return max(0, months)

@app.put("/revisoes/{rev_id}/itens/{item_id}", response_model=RevisaoItemOut)
def update_revisao_item(rev_id: int, item_id: int, payload: RevisaoItemUpdate, db: Session = Depends(get_db)):
    periodo = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail="Período de revisão não encontrado")

    item = db.query(RevisaoItemModel).filter(RevisaoItemModel.id == item_id).first()
    if not item or item.periodo_id != rev_id:
        raise HTTPException(status_code=404, detail="Item de revisão não encontrado para o período")

    # Se revisor_id informado, validar delegação ativa
    if payload.revisor_id is not None:
        deleg = db.query(RevisaoDelegacaoModel).filter(
            RevisaoDelegacaoModel.periodo_id == rev_id,
            RevisaoDelegacaoModel.ativo_id == item_id,
            RevisaoDelegacaoModel.revisor_id == payload.revisor_id,
            RevisaoDelegacaoModel.status == 'Ativo'
        ).first()
        if not deleg:
            raise HTTPException(status_code=403, detail="Revisor não autorizado para este item")

    data = payload.dict(exclude_unset=True)

    # Validações de negócio
    vida_revisada = data.get("vida_util_revisada")
    nova_data_fim = data.get("nova_data_fim")
    # Se informada nova_data_fim e não informada vida_util_revisada, calcular meses a partir do início definido no período
    if vida_revisada is None and nova_data_fim is not None:
        if not periodo.data_inicio_nova_vida_util:
            raise HTTPException(status_code=400, detail="Defina a Data de Início da Nova Vida Útil no cabeçalho do período")
        # calcular diferença em meses entre início e nova_data_fim
        start = periodo.data_inicio_nova_vida_util
        end = nova_data_fim
        # months diff (aproximação similar à função de front-end)
        months = (end.year - start.year) * 12 + (end.month - start.month)
        # Ajuste por dia do mês
        if end.day < start.day:
            months -= 1
        vida_revisada = max(0, months)
    if vida_revisada is not None:
        if vida_revisada <= 0:
            raise HTTPException(status_code=400, detail="Vida útil revisada deve ser maior que zero")
        # calcular data fim revisada com base na data de início prospectiva
        if not periodo.data_inicio_nova_vida_util:
            raise HTTPException(status_code=400, detail="Defina a Data de Início da Nova Vida Útil no cabeçalho do período")
        data_fim_rev = _add_months(periodo.data_inicio_nova_vida_util, vida_revisada)
        item.data_fim_revisada = data_fim_rev
        item.vida_util_revisada = vida_revisada
    # Se vida revisada não informada mas nova_data_fim informada, persistir nova_data_fim diretamente
    elif nova_data_fim is not None:
        item.data_fim_revisada = nova_data_fim

    # Calcular vida útil TOTAL original e revisada (do início da depreciação até o fim)
    original_total = item.vida_util_periodos or ( _months_diff(item.data_inicio_depreciacao, item.data_fim_depreciacao) if item.data_fim_depreciacao else 0 )
    revised_total = _months_diff(item.data_inicio_depreciacao, item.data_fim_revisada) if item.data_fim_revisada else None

    if "condicao_fisica" in data:
        cf = data.get("condicao_fisica")
        if cf is not None and cf not in {"Bom", "Regular", "Ruim"}:
            raise HTTPException(status_code=400, detail="Condição física inválida")
        item.condicao_fisica = cf

    # Incremento e motivo armazenados nos auxiliares
    if "incremento" in data:
        inc = data.get("incremento")
        if inc is not None and inc not in {"Acréscimo", "Decréscimo", "Manter"}:
            raise HTTPException(status_code=400, detail="Incremento inválido")

        # Regras de coerência entre incremento e vida revisada
        if inc == "Manter":
            # Não permitir alteração de vida útil total
            if revised_total is not None and original_total > 0 and revised_total != original_total:
                raise HTTPException(status_code=400, detail="Para 'Manter', não é permitido alterar a vida útil total")
        elif inc == "Decréscimo":
            # Exigir revisão (meses ou nova data) e ser menor que a original em termos de vida total
            if vida_revisada is None and nova_data_fim is None:
                raise HTTPException(status_code=400, detail="Para 'Decréscimo', informe a nova vida útil ou nova data fim")
            if revised_total is not None and original_total > 0 and revised_total >= original_total:
                raise HTTPException(status_code=400, detail="Para 'Decréscimo', a vida útil total revisada deve ser menor que a original")
        elif inc == "Acréscimo":
            # Exigir revisão (meses ou nova data) e ser maior que a original em termos de vida total
            if vida_revisada is None and nova_data_fim is None:
                raise HTTPException(status_code=400, detail="Para 'Acréscimo', informe a nova vida útil ou nova data fim")
            if revised_total is not None and original_total > 0 and revised_total <= original_total:
                raise HTTPException(status_code=400, detail="Para 'Acréscimo', a vida útil total revisada deve ser maior que a original")

        item.auxiliar2 = inc
    if "motivo" in data:
        mot = data.get("motivo")
        item.auxiliar3 = mot

    # Regras de justificativa
    justificativa = data.get("justificativa")
    now18m = _add_months(date.today(), 18)
    # se redução de vida útil TOTAL, exigir justificativa
    if revised_total is not None and original_total > 0 and revised_total < original_total:
        if not justificativa:
            raise HTTPException(status_code=400, detail="Justificativa obrigatória para redução de vida útil")
    # se nova data fim < 18 meses, justificativa obrigatória
    if item.data_fim_revisada and item.data_fim_revisada < now18m:
        if not justificativa:
            raise HTTPException(status_code=400, detail="Justificativa obrigatória para vida útil com vencimento < 18 meses")
    # alerta original <18 meses exige justificativa detalhada em caso de redução
    if item.data_fim_depreciacao and item.data_fim_depreciacao < now18m and revised_total is not None and original_total > 0 and revised_total < original_total:
        if not justificativa:
            raise HTTPException(status_code=400, detail="Justificativa obrigatória devido a risco (<18 meses)")
    # justificativa automática para itens não alterados
    if vida_revisada is None and not justificativa:
        item.justificativa = item.justificativa or "A vida útil está correta"
    elif justificativa is not None:
        item.justificativa = justificativa

    # Alterado se vida revisada difere da original
    if vida_revisada is not None:
        original_periodos = item.vida_util_periodos or 0
        # Se original é 0 (desconhecido), considerar como alteração
        item.alterado = (original_periodos == 0) or (vida_revisada != original_periodos)
        item.status = "Revisado" if item.alterado else (item.status or "Pendente")

    # Registrar autor se fornecido
    if payload.revisor_id is not None:
        item.criado_por = payload.revisor_id

    # Aprovação automática: incremento 'Manter' e vida útil restante > 18 meses
    try:
        inc_for_auto = data.get("incremento") if "incremento" in data else None
        effective_end = item.data_fim_revisada or item.data_fim_depreciacao
        # Apenas se período não estiver encerrado
        if inc_for_auto == "Manter" and effective_end and effective_end >= now18m and getattr(periodo, 'status', None) != 'Encerrado':
            item.status = 'Aprovado'
            # Primeiro commit da aprovação para garantir persistência mesmo se tabelas auxiliares não existirem
            db.commit()
            db.refresh(item)
            # Registrar histórico e auditoria (best-effort)
            try:
                revisada_total_hist = int(item.vida_util_revisada or (item.vida_util_periodos or 0))
                db.execute(sa.text(
                    """
                    INSERT INTO revisoes_historico (ativo_id, revisor_id, vida_util_revisada, acao, status)
                    VALUES (:ativo_id, :revisor_id, :vida_util_revisada, 'aprovado', 'aprovado')
                    """
                ), {
                    'ativo_id': item.id,
                    'revisor_id': payload.revisor_id,
                    'vida_util_revisada': revisada_total_hist,
                })
                db.execute(sa.text(
                    """
                    INSERT INTO auditoria_rvu (usuario_id, acao, entidade, entidade_id, detalhes)
                    VALUES (:usuario_id, 'auto_aprovar', 'revisao_item', :entidade_id, :detalhes)
                    """
                ), {
                    'usuario_id': payload.revisor_id,
                    'entidade_id': item.id,
                    'detalhes': 'Aprovacao automática: incremento=Manter e vida útil > 18 meses'
                })
                db.commit()
            except Exception:
                # Em caso de falha nas tabelas auxiliares, reverter apenas a transação auxiliar
                try:
                    db.rollback()
                except Exception:
                    pass
    except Exception:
        # Não bloquear atualização por falha na lógica opcional de aprovação automática
        pass

    db.commit()
    db.refresh(item)

    return RevisaoItemOut(
        id=item.id,
        periodo_id=item.periodo_id,
        numero_imobilizado=item.numero_imobilizado,
        sub_numero=item.sub_numero,
        descricao=item.descricao,
        data_inicio_depreciacao=item.data_inicio_depreciacao,
        data_fim_depreciacao=item.data_fim_depreciacao,
        data_fim_revisada=item.data_fim_revisada,
        valor_aquisicao=float(item.valor_aquisicao) if item.valor_aquisicao is not None else 0.0,
        depreciacao_acumulada=float(item.depreciacao_acumulada) if item.depreciacao_acumulada is not None else 0.0,
        valor_contabil=float(item.valor_contabil) if item.valor_contabil is not None else 0.0,
        centro_custo=item.centro_custo,
        classe=item.classe,
        descricao_classe=getattr(item, 'descricao_classe', None),
        conta_contabil=item.conta_contabil,
        descricao_conta_contabil=item.descricao_conta_contabil,
        vida_util_anos=item.vida_util_anos,
        vida_util_periodos=item.vida_util_periodos,
        vida_util_revisada=item.vida_util_revisada,
        condicao_fisica=item.condicao_fisica,
        justificativa=item.justificativa,
        alterado=item.alterado,
        auxiliar2=item.auxiliar2,
        auxiliar3=item.auxiliar3,
        status=item.status,
        criado_em=item.criado_em,
    )

# --- Revisão em Massa ---
class MassRevisionPayload(BaseModel):
    ativos_ids: List[int]
    incremento: Optional[str] = None  # Acréscimo | Decréscimo | Manter
    nova_vida_util_anos: Optional[int] = None
    nova_vida_util_meses: Optional[int] = None
    nova_data_fim: Optional[date] = None
    condicao_fisica: Optional[str] = None  # Bom | Regular | Ruim
    motivo: Optional[str] = None
    justificativa: Optional[str] = None

class MassRevisionItemResult(BaseModel):
    id: int
    numero_imobilizado: str
    changed: bool
    error: Optional[str] = None
    original_total_months: Optional[int] = None
    revised_total_months: Optional[int] = None
    original_end_date: Optional[date] = None
    revised_end_date: Optional[date] = None
    incremento: Optional[str] = None
    motivo: Optional[str] = None

class MassRevisionResponse(BaseModel):
    total: int
    updated: int
    skipped: int
    errors: List[str]
    results: List[MassRevisionItemResult]

@app.post("/revisoes/massa", response_model=MassRevisionResponse)
def aplicar_revisao_massa(payload: MassRevisionPayload, current_user: UsuarioModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.ativos_ids:
        raise HTTPException(status_code=400, detail="Lista de ativos vazia")

    inc = payload.incremento
    if inc is not None and inc not in {"Acréscimo", "Decréscimo", "Manter"}:
        raise HTTPException(status_code=400, detail="Incremento inválido")

    cf = payload.condicao_fisica
    if cf is not None and cf not in {"Bom", "Regular", "Ruim"}:
        raise HTTPException(status_code=400, detail="Condição física inválida")

    # total de meses informado diretamente
    meses_total_global: Optional[int] = None
    if payload.nova_vida_util_anos is not None or payload.nova_vida_util_meses is not None:
        anos = int(payload.nova_vida_util_anos or 0)
        meses = int(payload.nova_vida_util_meses or 0)
        meses_total_global = anos * 12 + meses
        if meses_total_global <= 0 and inc != "Manter":
            raise HTTPException(status_code=400, detail="Vida útil revisada deve ser maior que zero")

    results: List[MassRevisionItemResult] = []
    errors: List[str] = []
    updated = 0
    skipped = 0

    now18m = _add_months(date.today(), 18)

    for item_id in payload.ativos_ids:
        item = db.query(RevisaoItemModel).filter(RevisaoItemModel.id == item_id).first()
        if not item:
            msg = f"Item {item_id} não encontrado"
            errors.append(msg)
            results.append(MassRevisionItemResult(id=item_id, numero_imobilizado=str(item_id), changed=False, error=msg))
            skipped += 1
            continue

        # status bloqueado (se existir)
        if (item.status or "").lower() == "travado":
            msg = f"Item {item.id} está travado"
            errors.append(msg)
            results.append(MassRevisionItemResult(id=item.id, numero_imobilizado=item.numero_imobilizado, changed=False, error=msg))
            skipped += 1
            continue

        original_end = item.data_fim_depreciacao
        original_total = item.vida_util_periodos or ( _months_diff(item.data_inicio_depreciacao, item.data_fim_depreciacao) if item.data_fim_depreciacao else 0 )

        vida_revisada: Optional[int] = None
        nova_fim: Optional[date] = None

        # Se nova data fim foi informada, calcular meses com base na data de início do próprio ativo
        if payload.nova_data_fim is not None:
            end = payload.nova_data_fim
            start = item.data_inicio_depreciacao
            vida_revisada = _months_diff(start, end)
            nova_fim = end
        elif meses_total_global is not None:
            vida_revisada = meses_total_global
            nova_fim = _add_months(item.data_inicio_depreciacao, vida_revisada)

        # Coerência de incremento
        revised_total = _months_diff(item.data_inicio_depreciacao, nova_fim) if nova_fim else None

        try:
            if inc == "Manter":
                if revised_total is not None and original_total > 0 and revised_total != original_total:
                    raise HTTPException(status_code=400, detail="Para 'Manter', não é permitido alterar a vida útil total")
            elif inc == "Decréscimo":
                if vida_revisada is None and nova_fim is None:
                    raise HTTPException(status_code=400, detail="Para 'Decréscimo', informe a nova vida útil ou nova data fim")
                if revised_total is not None and original_total > 0 and revised_total >= original_total:
                    raise HTTPException(status_code=400, detail="Para 'Decréscimo', a vida útil total revisada deve ser menor que a original")
            elif inc == "Acréscimo":
                if vida_revisada is None and nova_fim is None:
                    raise HTTPException(status_code=400, detail="Para 'Acréscimo', informe a nova vida útil ou nova data fim")
                if revised_total is not None and original_total > 0 and revised_total <= original_total:
                    raise HTTPException(status_code=400, detail="Para 'Acréscimo', a vida útil total revisada deve ser maior que a original")

            # Validações gerais
            if vida_revisada is not None and vida_revisada <= 0 and inc != "Manter":
                raise HTTPException(status_code=400, detail="Vida útil revisada deve ser maior que zero")
            if nova_fim and nova_fim < now18m and inc != "Manter" and not (payload.justificativa or "").strip():
                raise HTTPException(status_code=400, detail="Justificativa obrigatória para vida útil com vencimento < 18 meses")

            # Persistir alterações
            if cf is not None:
                item.condicao_fisica = cf
            if inc is not None:
                item.auxiliar2 = inc
            if payload.motivo is not None:
                item.auxiliar3 = payload.motivo

            if vida_revisada is not None:
                item.vida_util_revisada = vida_revisada
                item.data_fim_revisada = nova_fim
                original_periodos = item.vida_util_periodos or 0
                item.alterado = (original_periodos == 0) or (vida_revisada != original_periodos)
                item.status = "Revisado" if item.alterado else (item.status or "Pendente")
            elif nova_fim is not None:
                # vida não informada, só data fim revisada
                item.data_fim_revisada = nova_fim

            if inc == "Manter" and not (payload.justificativa or "").strip():
                item.justificativa = item.justificativa or "A vida útil está correta"
            elif payload.justificativa is not None:
                item.justificativa = payload.justificativa

            if inc == "Manter":
                if not (item.status or '').strip() or (item.status or '').strip().lower() == 'pendente':
                    item.status = 'Revisado'
                effective_end = item.data_fim_revisada or item.data_fim_depreciacao
                per = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == item.periodo_id).first()
                if effective_end and effective_end >= now18m and getattr(per, 'status', None) != 'Encerrado':
                    item.status = 'Aprovado'

            updated += 1
            results.append(MassRevisionItemResult(
                id=item.id,
                numero_imobilizado=item.numero_imobilizado,
                changed=bool(item.alterado),
                original_total_months=original_total,
                revised_total_months=revised_total,
                original_end_date=original_end,
                revised_end_date=item.data_fim_revisada,
                incremento=item.auxiliar2,
                motivo=item.auxiliar3,
            ))
        except HTTPException as ex:
            msg = ex.detail
            errors.append(f"Item {item.id}: {msg}")
            results.append(MassRevisionItemResult(
                id=item.id,
                numero_imobilizado=item.numero_imobilizado,
                changed=False,
                error=msg,
                original_total_months=original_total,
                revised_total_months=revised_total,
                original_end_date=original_end,
                revised_end_date=nova_fim,
                incremento=inc,
                motivo=payload.motivo,
            ))
            skipped += 1

    db.commit()
    audit(db, usuario_id=current_user.id, acao="revisao_massa", entidade="revisao_item", entidade_id=None,
          detalhes=f"itens={updated}/{len(payload.ativos_ids)} inc={inc} motivo={payload.motivo}")

    return MassRevisionResponse(
        total=len(payload.ativos_ids),
        updated=updated,
        skipped=skipped,
        errors=errors,
        results=results,
    )

# -----------------------------
# Delegações de Revisão
# -----------------------------
class DelegacaoOut(BaseModel):
    id: int
    periodo_id: int
    ativo_id: int
    revisor_id: int
    atribuido_por: int
    data_atribuicao: datetime
    status: str
    numero_imobilizado: Optional[str] = None
    descricao: Optional[str] = None
    revisor_nome: Optional[str] = None
    classe: Optional[str] = None
    conta_contabil: Optional[str] = None
    centro_custo: Optional[str] = None
    valor_contabil: Optional[float] = None

class DelegacaoCreate(BaseModel):
    periodo_id: int
    ativo_id: int
    revisor_id: int
    atribuido_por: int

@app.get("/revisoes/delegacoes/{rev_id}", response_model=List[DelegacaoOut])
def list_revisao_delegacoes(rev_id: int, db: Session = Depends(get_db)):
    periodo = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == rev_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail="Período de revisão não encontrado")

    delegs = db.query(RevisaoDelegacaoModel).filter(RevisaoDelegacaoModel.periodo_id == rev_id).all()
    result: List[DelegacaoOut] = []
    for d in delegs:
        item = db.query(RevisaoItemModel).filter(RevisaoItemModel.id == d.ativo_id).first()
        revisor = db.query(UsuarioModel).filter(UsuarioModel.id == d.revisor_id).first()
        result.append(
            DelegacaoOut(
                id=d.id,
                periodo_id=d.periodo_id,
                ativo_id=d.ativo_id,
                revisor_id=d.revisor_id,
                atribuido_por=d.atribuido_por,
                data_atribuicao=d.data_atribuicao,
                status=d.status,
                numero_imobilizado=item.numero_imobilizado if item else None,
                descricao=item.descricao if item else None,
                revisor_nome=revisor.nome_completo if revisor else None,
                classe=item.classe if item else None,
                conta_contabil=item.conta_contabil if item else None,
                centro_custo=item.centro_custo if item else None,
                valor_contabil=(float(item.valor_contabil) if (hasattr(item, 'valor_contabil') and item.valor_contabil is not None) else None),
            )
        )
    return result

@app.post("/revisoes/delegacoes", response_model=DelegacaoOut)
def create_revisao_delegacao(payload: DelegacaoCreate, db: Session = Depends(get_db)):
    periodo = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == payload.periodo_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail="Período de revisão não encontrado")

    if periodo.status in ['Encerrado', 'Concluído', 'Fechado']:
        raise HTTPException(status_code=400, detail="Não é permitido criar delegações em um período encerrado.")

    item = db.query(RevisaoItemModel).filter(RevisaoItemModel.id == payload.ativo_id).first()
    if not item or item.periodo_id != payload.periodo_id:
        raise HTTPException(status_code=400, detail="Ativo inválido para o período")

    revisor = db.query(UsuarioModel).filter(UsuarioModel.id == payload.revisor_id).first()
    if not revisor:
        raise HTTPException(status_code=400, detail="Revisor inválido")

    # Regra de negócio: delegar o grupo inteiro (ativo principal e incorporações)
    group_items = (
        db.query(RevisaoItemModel)
        .filter(
            RevisaoItemModel.periodo_id == payload.periodo_id,
            RevisaoItemModel.numero_imobilizado == item.numero_imobilizado,
        )
        .all()
    )
    group_ids = [gi.id for gi in group_items]

    # Delegações ativas já existentes para o grupo
    existing_group_delegs = (
        db.query(RevisaoDelegacaoModel)
        .filter(
            RevisaoDelegacaoModel.periodo_id == payload.periodo_id,
            RevisaoDelegacaoModel.ativo_id.in_(group_ids),
            RevisaoDelegacaoModel.status == 'Ativo',
        )
        .all()
    )

    # Se já houver delegações com revisor diferente, impedir separação
    existing_revisors = {d.revisor_id for d in existing_group_delegs}
    if len(existing_revisors) > 0 and existing_revisors != {payload.revisor_id}:
        raise HTTPException(
            status_code=400,
            detail="Regra: ativo principal e incorporações devem ser revisados pelo mesmo revisor"
        )

    # Impedir duplicidade para o item solicitado
    if any(d.ativo_id == payload.ativo_id for d in existing_group_delegs):
        raise HTTPException(status_code=400, detail="Ativo já delegado")

    # Criar delegações para todo o grupo que ainda não possui
    created_for_requested = None
    for gi in group_items:
        if any(d.ativo_id == gi.id for d in existing_group_delegs):
            continue
        d = RevisaoDelegacaoModel(
            periodo_id=payload.periodo_id,
            ativo_id=gi.id,
            revisor_id=payload.revisor_id,
            atribuido_por=payload.atribuido_por,
            status='Ativo'
        )
        db.add(d)
        db.flush()  # obter id sem commit final
        if gi.id == payload.ativo_id:
            created_for_requested = d

    db.commit()
    if created_for_requested:
        d = created_for_requested
    else:
        # por segurança, recuperar delegação do item solicitado
        d = db.query(RevisaoDelegacaoModel).filter(
            RevisaoDelegacaoModel.periodo_id == payload.periodo_id,
            RevisaoDelegacaoModel.ativo_id == payload.ativo_id,
            RevisaoDelegacaoModel.status == 'Ativo'
        ).first()

    return DelegacaoOut(
        id=d.id,
        periodo_id=d.periodo_id,
        ativo_id=d.ativo_id,
        revisor_id=d.revisor_id,
        atribuido_por=d.atribuido_por,
        data_atribuicao=d.data_atribuicao,
        status=d.status,
        numero_imobilizado=item.numero_imobilizado,
        descricao=item.descricao,
        revisor_nome=revisor.nome_completo,
        classe=item.classe,
        conta_contabil=item.conta_contabil,
        centro_custo=item.centro_custo,
        valor_contabil=(float(item.valor_contabil) if item.valor_contabil is not None else None),
    )

@app.delete("/revisoes/delegacoes/{delegacao_id}")
def delete_revisao_delegacao(delegacao_id: int, db: Session = Depends(get_db)):
    d = db.query(RevisaoDelegacaoModel).filter(RevisaoDelegacaoModel.id == delegacao_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Delegação não encontrada")

    periodo = db.query(RevisaoPeriodoModel).filter(RevisaoPeriodoModel.id == d.periodo_id).first()
    if periodo and periodo.status in ['Encerrado', 'Concluído', 'Fechado']:
        raise HTTPException(status_code=400, detail="Não é permitido remover delegações em um período encerrado.")

    # Regra: não separar ativo principal (Sub nº 0) das incorporações (Sub nº > 0)
    # Ao remover uma delegação, remover todo o grupo do mesmo imobilizado neste período.
    item = db.query(RevisaoItemModel).filter(RevisaoItemModel.id == d.ativo_id).first()
    if not item:
        # fallback: se não localizar o item, remove somente o registro atual
        db.delete(d)
        db.commit()
        return {"deleted": True, "group": 0}

    group_items = (
        db.query(RevisaoItemModel)
        .filter(
            RevisaoItemModel.periodo_id == d.periodo_id,
            RevisaoItemModel.numero_imobilizado == item.numero_imobilizado,
        )
        .all()
    )
    group_ids = [gi.id for gi in group_items]

    # Remover todas as delegações ativas do grupo nesse período
    group_delegs = (
        db.query(RevisaoDelegacaoModel)
        .filter(
            RevisaoDelegacaoModel.periodo_id == d.periodo_id,
            RevisaoDelegacaoModel.ativo_id.in_(group_ids),
        )
        .all()
    )
    for gd in group_delegs:
        db.delete(gd)

    db.commit()
    return {"deleted": True, "group": len(group_delegs)}
